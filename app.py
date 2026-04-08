# -*- coding: utf-8 -*-
"""
Analizador de Revistas OJS v5
"""

import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import urllib3
import time
import re
import io
import json
from collections import Counter
from itertools import combinations

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np

import plotly.graph_objects as go
import plotly.express as px

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA, LatentDirichletAllocation
import networkx as nx
from wordcloud import WordCloud
import nltk
from nltk.corpus import stopwords

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm as rcm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image as RLImage, PageBreak, HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}

# ─────────────────────────────────────────
# PALETA DE COLORES AZUL PERSONALIZADA
# ─────────────────────────────────────────
PALETTE = {
    "50":  "#E8F4FB",
    "100": "#BCD9F0",
    "200": "#82B8E2",
    "300": "#048ABF",
    "400": "#04B2D9",
    "500": "#0367A6",
    "600": "#055BA6",
    "700": "#04507A",
    "800": "#033A5A",
    "900": "#02243A",
    "950": "#038C7F",   # Teal accent
    "teal": "#038C7F",
    "teal_light": "#04B2D9",
}

BLUE_SCALE = [
    [0.0,  PALETTE["50"]],
    [0.25, PALETTE["300"]],
    [0.5,  PALETTE["500"]],
    [0.75, PALETTE["600"]],
    [1.0,  PALETTE["800"]],
]

def hex_to_rgb(hex_color):
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def palette_color(key, alpha=1.0):
    r, g, b = hex_to_rgb(PALETTE[key])
    if alpha < 1.0:
        return f"rgba({r},{g},{b},{alpha})"
    return PALETTE[key]

# ─────────────────────────────────────────
# STOPWORDS EXTENDIDAS PARA CLUSTERS
# ─────────────────────────────────────────
STOPWORDS_ACADEMICAS = {
    # Palabras académicas genéricas (español)
    "articulo", "artículo", "libro", "capítulo", "capitulo", "revista",
    "journal", "paper", "estudio", "estudios", "análisis", "analisis",
    "investigación", "investigacion", "trabajo", "trabajos", "resultado",
    "resultados", "conclusión", "conclusion", "conclusiones", "método",
    "metodo", "métodos", "metodos", "objetivo", "objetivos", "propuesta",
    "propósito", "propuesto", "revisión", "revision", "enfoque", "enfoques",
    "caso", "casos", "datos", "base", "bases", "sistema", "sistemas",
    "proceso", "procesos", "modelo", "modelos", "análise", "presente",
    "mediante", "través", "través", "partir", "través", "hacia",
    "también", "asimismo", "además", "sin", "embargo", "aunque",
    "donde", "cual", "cuales", "cuyo", "cuyos", "cuya", "cuyas",
    "ante", "bajo", "sobre", "para", "este", "esta", "estos", "estas",
    "aquel", "aquella", "aquellos", "aquellas", "dicho", "dichos",
    "dicha", "dichas", "mismo", "mismos", "misma", "mismas",
    "bien", "tal", "tales", "cada", "todo", "todos", "toda", "todas",
    "otro", "otros", "otra", "otras", "nuevo", "nuevos", "nueva", "nuevas",
    "gran", "grandes", "importante", "importantes", "general", "generales",
    "diferente", "diferentes", "específico", "específicos", "específica",
    "principal", "principales", "mayor", "menores", "menor", "mayores",
    "uso", "usos", "tipo", "tipos", "forma", "formas", "nivel", "niveles",
    "factor", "factores", "aspecto", "aspectos", "elemento", "elementos",
    "parte", "partes", "área", "áreas", "campo", "campos", "tema", "temas",
    "información", "conocimiento", "perspectiva", "perspectivas",
    "contexto", "contextos", "situación", "situaciones",
    # Palabras académicas genéricas (inglés)
    "article", "study", "studies", "research", "paper", "work", "analysis",
    "method", "methods", "result", "results", "conclusion", "approach",
    "approaches", "review", "reviews", "model", "models", "data", "based",
    "using", "used", "use", "new", "different", "important", "general",
    "present", "proposed", "case", "cases", "system", "systems",
    "process", "processes", "however", "also", "although", "well",
    "may", "can", "two", "three", "four", "five", "first", "second",
    "thus", "therefore", "whereas", "among", "within", "between",
    # OJS / publicación
    "vol", "núm", "num", "pp", "doi", "issn", "http", "https", "www",
    "resumen", "abstract", "keywords", "palabras", "clave",
    # Números y símbolos
    "año", "años", "mes", "meses", "figura", "tabla", "gráfico",
}

# ─────────────────────────────────────────
# GEOCODIFICACIÓN
# ─────────────────────────────────────────
COUNTRY_COORDS = {
    "mexico": (23.6345, -102.5528), "méxico": (23.6345, -102.5528),
    "argentina": (-38.4161, -63.6167), "brasil": (-14.235, -51.9253),
    "brazil": (-14.235, -51.9253), "colombia": (4.5709, -74.2973),
    "chile": (-35.6751, -71.5430), "perú": (-9.19, -75.0152),
    "peru": (-9.19, -75.0152), "venezuela": (6.4238, -66.5897),
    "ecuador": (-1.8312, -78.1834), "bolivia": (-16.2902, -63.5887),
    "uruguay": (-32.5228, -55.7658), "paraguay": (-23.4425, -58.4438),
    "costa rica": (9.7489, -83.7534), "cuba": (21.5218, -77.7812),
    "españa": (40.4637, -3.7492), "spain": (40.4637, -3.7492),
    "united states": (37.0902, -95.7129), "estados unidos": (37.0902, -95.7129),
    "usa": (37.0902, -95.7129), "canada": (56.1304, -106.3468),
    "united kingdom": (55.3781, -3.4360), "reino unido": (55.3781, -3.4360),
    "france": (46.2276, 2.2137), "francia": (46.2276, 2.2137),
    "germany": (51.1657, 10.4515), "alemania": (51.1657, 10.4515),
    "italy": (41.8719, 12.5674), "italia": (41.8719, 12.5674),
    "portugal": (39.3999, -8.2245),
    "japan": (36.2048, 138.2529), "japón": (36.2048, 138.2529),
    "china": (35.8617, 104.1954), "india": (20.5937, 78.9629),
    "australia": (-25.2744, 133.7751), "russia": (61.5240, 105.3188),
    "rusia": (61.5240, 105.3188), "south africa": (-30.5595, 22.9375),
    "sudáfrica": (-30.5595, 22.9375), "nigeria": (9.0820, 8.6753),
    "kenya": (-0.0236, 37.9062), "egypt": (26.8206, 30.8025),
    "egipto": (26.8206, 30.8025), "netherlands": (52.1326, 5.2913),
    "holanda": (52.1326, 5.2913), "sweden": (60.1282, 18.6435),
    "suecia": (60.1282, 18.6435), "switzerland": (46.8182, 8.2275),
    "suiza": (46.8182, 8.2275), "korea": (35.9078, 127.7669),
    "corea": (35.9078, 127.7669), "panama": (8.5380, -80.7821),
    "panamá": (8.5380, -80.7821), "guatemala": (15.7835, -90.2308),
    "honduras": (15.2000, -86.2419), "el salvador": (13.7942, -88.8965),
    "nicaragua": (12.8654, -85.2072), "dominican republic": (18.7357, -70.1627),
    "república dominicana": (18.7357, -70.1627),
}

PAIS_PALABRAS_CLAVE = {
    "unam": "mexico", "ipn": "mexico", "uam": "mexico", "cinvestav": "mexico",
    "imss": "mexico", "inifap": "mexico", "colmex": "mexico",
    "uba": "argentina", "conicet": "argentina", "unlp": "argentina",
    "usp": "brasil", "unicamp": "brasil", "fiocruz": "brasil",
    "unal": "colombia", "udea": "colombia", "andes": "colombia",
    "uc": "chile", "uchile": "chile", "usach": "chile",
    "pucp": "perú", "unmsm": "perú", "cayetano": "perú",
    "ucv": "venezuela", "ivic": "venezuela",
    "uce": "ecuador", "espol": "ecuador", "flacso": "ecuador",
    "umsa": "bolivia", "umss": "bolivia",
    "csic": "españa", "ub.edu": "españa",
    "mit": "united states", "stanford": "united states", "harvard": "united states",
    "oxford": "united kingdom", "cambridge": "united kingdom",
}

def geocodificar_pais(afiliacion_texto):
    if not afiliacion_texto:
        return None, None, None
    texto = afiliacion_texto.lower().strip()
    for kw, pais in PAIS_PALABRAS_CLAVE.items():
        if kw in texto:
            coords = COUNTRY_COORDS.get(pais)
            if coords:
                return pais.title(), coords[0], coords[1]
    for pais, coords in COUNTRY_COORDS.items():
        if pais in texto:
            return pais.title(), coords[0], coords[1]
    return None, None, None


# ─────────────────────────────────────────
# NLTK setup
# ─────────────────────────────────────────
@st.cache_resource
def load_stopwords():
    try:
        nltk.download("stopwords", quiet=True)
        sw = set(stopwords.words("spanish"))
        sw.update(stopwords.words("english"))
        sw.update([
            "resumen", "abstract", "article", "revista", "journal",
            "vol", "núm", "num", "pp", "doi", "issn", "http", "https",
            "www", "artículo", "estudio", "análisis", "trabajo", "investigación",
            "este", "esta", "estos", "estas", "también", "así", "mediante",
            "través", "partir", "cada", "cual", "cuales", "sino", "donde",
        ])
        sw.update(STOPWORDS_ACADEMICAS)
        return sw
    except Exception:
        return set(STOPWORDS_ACADEMICAS)


# ─────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="Analizador OJS v5",
    page_icon="🔵",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────
# CUSTOM CSS — PALETA AZUL, TEXTO NEGRO
# ─────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

:root {{
  --c50:  {PALETTE["50"]};
  --c100: {PALETTE["100"]};
  --c200: {PALETTE["200"]};
  --c300: {PALETTE["300"]};
  --c400: {PALETTE["400"]};
  --c500: {PALETTE["500"]};
  --c600: {PALETTE["600"]};
  --c700: {PALETTE["700"]};
  --c800: {PALETTE["800"]};
  --c900: {PALETTE["900"]};
  --teal: {PALETTE["teal"]};
  --teal-light: {PALETTE["teal_light"]};
  --text: #0a0a0a;
}}

html, body, [class*="css"] {{
  font-family: 'Space Grotesk', sans-serif;
  color: var(--text);
}}

.main-header {{
  background: linear-gradient(135deg, {PALETTE["800"]} 0%, {PALETTE["600"]} 50%, {PALETTE["500"]} 100%);
  padding: 2.5rem 2rem; border-radius: 16px; margin-bottom: 2rem;
  text-align: center; color: white;
  border: 1px solid rgba(4,178,217,0.3);
  box-shadow: 0 8px 40px rgba(3,58,90,0.4);
  position: relative; overflow: hidden;
}}
.main-header::before {{
  content: '';
  position: absolute; top: 0; left: 0; right: 0; bottom: 0;
  background: radial-gradient(ellipse at 20% 50%, rgba(4,178,217,0.12) 0%, transparent 60%),
              radial-gradient(ellipse at 80% 20%, rgba(4,138,191,0.08) 0%, transparent 50%);
  pointer-events: none;
}}
.main-header h1 {{
  font-size: 2.4rem; margin: 0; font-weight: 700; letter-spacing: -0.5px;
  background: linear-gradient(90deg, #ffffff, {PALETTE["400"]}, #ffffff);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}}
.main-header p {{ font-size: 0.95rem; margin: 0.6rem 0 0; opacity: 0.8; font-weight: 300; color: #E8F4FB; }}
.main-header .version-badge {{
  display: inline-block; background: rgba(4,178,217,0.2);
  border: 1px solid rgba(4,178,217,0.5); border-radius: 20px;
  padding: 2px 14px; font-size: 0.75rem; margin-top: 0.5rem;
  font-family: 'JetBrains Mono', monospace; color: {PALETTE["400"]};
}}

.metric-card {{
  background: linear-gradient(145deg, #ffffff, {PALETTE["50"]});
  border-left: 4px solid {PALETTE["600"]}; padding: 1rem 1.2rem;
  border-radius: 10px; box-shadow: 0 2px 12px rgba(5,91,166,0.1);
  margin-bottom: 1rem; transition: transform 0.15s ease, box-shadow 0.15s ease;
  color: var(--text);
}}
.metric-card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 20px rgba(5,91,166,0.18); }}
.metric-card h3 {{ margin: 0; font-size: 2.2rem; color: {PALETTE["600"]}; font-weight: 700; }}
.metric-card p  {{ margin: 0; color: #374151; font-size: 0.85rem; }}
.metric-card.accent h3 {{ color: {PALETTE["300"]}; }}
.metric-card.teal h3 {{ color: {PALETTE["teal"]}; }}
.metric-card.dark h3 {{ color: {PALETTE["800"]}; }}

.section-header {{
  background: linear-gradient(90deg, {PALETTE["50"]}, {PALETTE["100"]});
  border-left: 4px solid {PALETTE["600"]}; padding: 0.7rem 1.2rem;
  border-radius: 6px; margin: 1.8rem 0 1rem; font-weight: 600;
  color: {PALETTE["800"]}; font-size: 1.05rem; letter-spacing: -0.2px;
}}

.period-filter {{
  background: linear-gradient(135deg, {PALETTE["800"]}, {PALETTE["700"]});
  border: 1px solid rgba(4,178,217,0.3); border-radius: 12px;
  padding: 1.2rem 1.5rem; margin: 1rem 0;
  box-shadow: 0 4px 16px rgba(3,58,90,0.3);
}}
.period-filter h4 {{
  color: {PALETTE["400"]}; margin: 0 0 0.8rem;
  font-family: 'JetBrains Mono', monospace; font-size: 0.9rem;
  letter-spacing: 0.05em; text-transform: uppercase;
}}

.info-box {{
  background: linear-gradient(135deg, {PALETTE["50"]}, #ffffff);
  border: 1px solid {PALETTE["200"]}; border-radius: 10px;
  padding: 1.2rem 1.5rem; margin: 0.8rem 0; color: var(--text);
}}
.info-box h4 {{ color: {PALETTE["700"]}; margin: 0 0 0.5rem; }}
.info-box p, .info-box li {{ color: #1f2937; font-size: 0.9rem; }}

.explanation-box {{
  background: linear-gradient(135deg, #ffffff, {PALETTE["50"]});
  border: 1px solid {PALETTE["200"]}; border-radius: 12px;
  padding: 1.4rem 1.6rem; margin: 1rem 0; color: var(--text);
  border-top: 3px solid {PALETTE["600"]};
}}
.explanation-box h4 {{ color: {PALETTE["700"]}; margin: 0 0 0.8rem; font-size: 1rem; }}
.explanation-box .formula {{
  background: {PALETTE["800"]}; color: {PALETTE["400"]};
  padding: 0.6rem 1rem; border-radius: 6px;
  font-family: 'JetBrains Mono', monospace; font-size: 0.85rem;
  margin: 0.6rem 0; display: block;
}}

.network-controls {{
  background: {PALETTE["800"]}; border: 1px solid rgba(4,178,217,0.3);
  border-radius: 10px; padding: 0.8rem 1rem; margin-bottom: 1rem;
  color: {PALETTE["200"]}; font-size: 0.88rem;
}}

.stProgress > div > div {{ background-color: {PALETTE["500"]}; }}
.stTabs [data-baseweb="tab-list"] {{ gap: 4px; }}
.stTabs [data-baseweb="tab"] {{
  background: {PALETTE["50"]}; border-radius: 8px 8px 0 0;
  border: 1px solid {PALETTE["200"]}; color: {PALETTE["800"]};
  font-weight: 500;
}}
.stTabs [aria-selected="true"] {{
  background: {PALETTE["600"]} !important; color: white !important;
}}

.stButton > button[kind="primary"] {{
  background: linear-gradient(135deg, {PALETTE["600"]}, {PALETTE["500"]});
  border: none; border-radius: 8px; font-weight: 600;
  color: white !important;
  transition: all 0.2s ease;
}}
.stButton > button[kind="primary"]:hover {{
  background: linear-gradient(135deg, {PALETTE["500"]}, {PALETTE["300"]});
  box-shadow: 0 4px 16px rgba(5,91,166,0.4);
}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1> Analizador de Revistas OJS</h1>
    <p>Extracción avanzada · Análisis de contenido · Filtrado por período · Mapa geográfico · Red interactiva con ORCID</p>
    <span class="version-badge">v5.0 · OJS 2.x / 3.x · Paleta azul · Excel + PDF mejorado</span>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════
# ── SCRAPING ──────────────────────────────────────────
# ═══════════════════════════════════════════════════════

def parsear_issue_metadata(texto, href):
    resultado = {
        "titulo_original": texto, "issue_url": href, "tipo": "Artículo",
        "volumen": "", "numero": "", "anio": "", "issue": texto,
    }
    m_anio = re.search(r"\b(19|20)\d{2}\b", texto)
    if m_anio:
        resultado["anio"] = m_anio.group()
    m_vol = re.search(r"[Vv]ol(?:umen|ume|\.)?\.?\s*(\d+)", texto)
    if m_vol:
        resultado["volumen"] = m_vol.group(1)
    m_num = re.search(r"[Nn]ú?[mn](?:ero)?\.?\s*[°º]?\s*(\d+)", texto)
    if not m_num:
        m_num = re.search(r"[Nn]o?\.?\s*(\d+)", texto)
    if not m_num:
        m_num = re.search(r"#\s*(\d+)", texto)
    if m_num:
        resultado["numero"] = m_num.group(1)
    tl = texto.lower()
    if any(x in tl for x in ["suplemento", "supplement", "supl"]):
        resultado["tipo"] = "Suplemento"
    elif any(x in tl for x in ["especial", "special", "monográf"]):
        resultado["tipo"] = "Especial"
    elif any(x in tl for x in ["número", "numero", "issue", "no.", "núm"]):
        resultado["tipo"] = "Número"
    elif any(x in tl for x in ["volumen", "volume", "vol"]):
        resultado["tipo"] = "Volumen"
    elif re.search(r"\b(19|20)\d{2}\b", texto):
        resultado["tipo"] = "Anual"
    return resultado


def extraer_issues_todas_paginas(url_base):
    issues_vistos = set()
    todos_issues = []
    error = None
    url_limpia = re.sub(r"[?&]page=\d+", "", url_base).rstrip("?&")
    paginas_visitadas = set()
    urls_a_visitar = [url_limpia]

    while urls_a_visitar:
        url_actual = urls_a_visitar.pop(0)
        if url_actual in paginas_visitadas:
            continue
        paginas_visitadas.add(url_actual)
        try:
            resp = requests.get(url_actual, headers=HEADERS, verify=False, timeout=15)
            resp.raise_for_status()
        except Exception as e:
            error = str(e)
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        base_url = "/".join(url_base.split("/")[:3])
        nuevos = 0

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/issue/view/" not in href:
                continue
            texto = a.get_text(strip=True)
            if not texto or len(texto) < 3:
                continue
            if href.startswith("/"):
                href = base_url + href
            if href not in issues_vistos:
                issues_vistos.add(href)
                meta = parsear_issue_metadata(texto, href)
                todos_issues.append(meta)
                nuevos += 1

        for a in soup.find_all("a", href=True):
            href_pag = a["href"]
            texto_a = a.get_text(strip=True).lower()
            es_siguiente = any(x in texto_a for x in ["siguiente", "next", "›", "»", ">"])
            tiene_page = "page=" in href_pag or "/issue/archive/" in href_pag
            if (es_siguiente or tiene_page) and href_pag not in paginas_visitadas:
                if href_pag.startswith("/"):
                    href_pag = base_url + href_pag
                if href_pag not in paginas_visitadas and "issue" in href_pag.lower():
                    urls_a_visitar.append(href_pag)

        if nuevos > 0 and url_actual == url_limpia:
            page_links = soup.find_all("a", href=re.compile(r"[?&]page=\d+"))
            for pl in page_links:
                ph = pl["href"]
                if ph.startswith("/"):
                    ph = base_url + ph
                if ph not in paginas_visitadas:
                    urls_a_visitar.append(ph)
            for page_n in range(2, 50):
                sep = "&" if "?" in url_limpia else "?"
                url_pag = f"{url_limpia}{sep}page={page_n}"
                if url_pag not in paginas_visitadas:
                    try:
                        r_check = requests.get(url_pag, headers=HEADERS, verify=False, timeout=10)
                        check_soup = BeautifulSoup(r_check.text, "html.parser")
                        tiene_issues = any(
                            "/issue/view/" in a.get("href", "")
                            for a in check_soup.find_all("a", href=True)
                        )
                        if tiene_issues:
                            urls_a_visitar.append(url_pag)
                        else:
                            break
                    except Exception:
                        break
        time.sleep(0.3)

    df = (pd.DataFrame(todos_issues).drop_duplicates(subset=["issue_url"])
          if todos_issues else pd.DataFrame())
    return df, error


def extraer_metadatos_metatags(soup):
    meta = {
        "doi": "", "titulo_meta": "", "abstract_meta": "",
        "keywords_meta": "", "autores_meta": [],
        "afiliaciones_meta": [], "pais_meta": "",
        "fecha_pub": "", "volumen_meta": "", "numero_meta": "",
        "pagina_inicio": "", "pagina_fin": "",
        "issn_meta": "", "journal_meta": "",
        "orcids_meta": [],
    }
    for tag in soup.find_all("meta"):
        name = (tag.get("name") or tag.get("property") or "").lower().strip()
        cont = (tag.get("content") or "").strip()
        if not cont:
            continue
        if name in ("citation_doi", "dc.identifier", "dc.identifier.doi"):
            if "10." in cont:
                meta["doi"] = cont.replace("https://doi.org/", "").replace("http://doi.org/", "")
        elif name == "citation_title":
            meta["titulo_meta"] = cont
        elif name in ("citation_abstract", "dc.description", "description"):
            if len(cont) > len(meta["abstract_meta"]):
                meta["abstract_meta"] = cont
        elif name in ("citation_keywords", "keywords", "dc.subject"):
            meta["keywords_meta"] = cont
        elif name in ("citation_author", "dc.creator", "dc.contributor"):
            if cont not in meta["autores_meta"]:
                meta["autores_meta"].append(cont)
        elif name == "citation_author_institution":
            if cont not in meta["afiliaciones_meta"]:
                meta["afiliaciones_meta"].append(cont)
        elif name == "citation_author_orcid":
            if cont not in meta["orcids_meta"]:
                meta["orcids_meta"].append(cont)
        elif name in ("citation_publication_date", "citation_date", "dc.date"):
            meta["fecha_pub"] = cont[:10]
        elif name == "citation_volume":
            meta["volumen_meta"] = cont
        elif name == "citation_issue":
            meta["numero_meta"] = cont
        elif name == "citation_firstpage":
            meta["pagina_inicio"] = cont
        elif name == "citation_lastpage":
            meta["pagina_fin"] = cont
        elif name in ("citation_issn", "citation_eissn"):
            meta["issn_meta"] = cont
        elif name == "citation_journal_title":
            meta["journal_meta"] = cont
    return meta


def extraer_autores_con_afiliacion(soup):
    autores = []
    autores_bloque = (
        soup.find("ul", class_=re.compile(r"authors", re.I)) or
        soup.find("div", class_=re.compile(r"authors", re.I)) or
        soup.find("section", class_=re.compile(r"authors", re.I))
    )
    if autores_bloque:
        items = autores_bloque.find_all(
            ["li", "div", "span"],
            class_=re.compile(r"author|contributor", re.I)
        )
        if not items:
            items = [autores_bloque]
        for item in items:
            nombre = ""
            afiliacion = ""
            orcid_url = ""
            for cls in ["name", "author-name", "contrib-name", "given-names"]:
                n_tag = item.find(class_=re.compile(cls, re.I))
                if n_tag:
                    nombre = n_tag.get_text(strip=True)
                    break
            if not nombre:
                n_tag = item.find(["strong", "b", "a"])
                if n_tag:
                    nombre = n_tag.get_text(strip=True)
            for cls in ["affiliation", "institution", "afiliacion", "aff"]:
                a_tag = item.find(class_=re.compile(cls, re.I))
                if a_tag:
                    afiliacion = a_tag.get_text(strip=True)
                    break
            for a_tag in item.find_all("a", href=True):
                if "orcid.org" in a_tag["href"]:
                    orcid_url = a_tag["href"]
                    break
            if nombre:
                pais, lat, lon = geocodificar_pais(afiliacion)
                autores.append({
                    "nombre": nombre, "afiliacion": afiliacion,
                    "orcid": orcid_url, "pais": pais or "",
                    "lat": lat, "lon": lon,
                })
    if not autores:
        for contenedor_cls in ["article-details", "articleDetails", "main-content", "pkp_structure_main"]:
            contenedor = soup.find(class_=re.compile(contenedor_cls, re.I))
            if contenedor:
                break
        else:
            contenedor = soup
        for tag in contenedor.find_all(class_=re.compile(r"author|contributor", re.I)):
            texto = tag.get_text(strip=True)
            if texto and 3 < len(texto) < 100:
                orcid_url = ""
                for a_tag in tag.find_all("a", href=True):
                    if "orcid.org" in a_tag["href"]:
                        orcid_url = a_tag["href"]
                        break
                siguiente = tag.find_next_sibling()
                afiliacion = ""
                if siguiente:
                    sib_text = siguiente.get_text(strip=True)
                    if len(sib_text) < 200:
                        afiliacion = sib_text
                pais, lat, lon = geocodificar_pais(afiliacion)
                autores.append({
                    "nombre": texto, "afiliacion": afiliacion,
                    "orcid": orcid_url, "pais": pais or "",
                    "lat": lat, "lon": lon,
                })
    return autores


def extraer_abstract_mejorado(soup, meta_tags):
    if meta_tags.get("abstract_meta") and len(meta_tags["abstract_meta"]) > 40:
        return meta_tags["abstract_meta"]
    selectores = [
        {"class_": "abstract"}, {"class_": "article-abstract"},
        {"class_": re.compile(r"abstract", re.I)}, {"id": "abstract"},
        {"class_": "abstractSection"}, {"class_": "item abstract"},
    ]
    for sel in selectores:
        try:
            tag = soup.find(**sel)
            if tag:
                texto = tag.get_text(separator=" ", strip=True)
                for prefix in ["Abstract", "Resumen", "RESUMEN", "ABSTRACT", "Resumo"]:
                    texto = re.sub(rf"^{prefix}[\s:]+", "", texto, flags=re.I).strip()
                if len(texto) > 50:
                    return texto
        except Exception:
            continue
    return ""


def extraer_keywords_mejorado(soup, meta_tags):
    if meta_tags.get("keywords_meta"):
        return meta_tags["keywords_meta"]
    selectores = [
        {"class_": "keywords"}, {"class_": "article-keywords"},
        {"class_": re.compile(r"keyword", re.I)}, {"id": "keywords"},
        {"class_": "item keywords"},
    ]
    for sel in selectores:
        try:
            tag = soup.find(**sel)
            if tag:
                texto = tag.get_text(separator=", ", strip=True)
                for prefix in ["Palabras clave:", "Keywords:", "Palabras Clave:", "Key words:"]:
                    texto = re.sub(re.escape(prefix), "", texto, flags=re.I).strip()
                if texto and len(texto) > 3:
                    return texto[:500]
        except Exception:
            continue
    return ""


def extraer_doi_mejorado(soup, meta_tags, url_articulo):
    if meta_tags.get("doi"):
        doi = meta_tags["doi"]
        return doi if doi.startswith("10.") else doi.replace("https://doi.org/", "")
    patrones_doi = [r"10\.\d{4,9}/[^\s\"<>'\]]+", r"doi\.org/(10\.[^\s\"<>'\]]+)"]
    texto_pagina = soup.get_text(" ")
    for pat in patrones_doi:
        m = re.search(pat, texto_pagina)
        if m:
            doi = m.group()
            doi = re.sub(r"^.*?(10\.)", r"10.", doi)
            return doi.rstrip(".,;)")
    for a in soup.find_all("a", href=True):
        if "doi.org" in a["href"] and "10." in a["href"]:
            m = re.search(r"10\.\d{4,9}/[^\s\"<>]+", a["href"])
            if m:
                return m.group().rstrip(".,;)")
    m = re.search(r"10\.\d{4,9}/[^\s\"<>]+", url_articulo)
    if m:
        return m.group()
    return ""


def extraer_orcids_mejorado(soup, meta_tags, autores_lista):
    orcid_map = {}
    if meta_tags.get("orcids_meta") and meta_tags.get("autores_meta"):
        for i, orcid_url in enumerate(meta_tags["orcids_meta"]):
            if i < len(meta_tags["autores_meta"]):
                orcid_map[meta_tags["autores_meta"][i]] = orcid_url
            elif i < len(autores_lista):
                orcid_map[autores_lista[i]] = orcid_url
    orcid_pattern = re.compile(r"orcid\.org/(\d{4}-\d{4}-\d{4}-\d{3}[\dX])", re.I)
    for a in soup.find_all("a", href=True):
        if "orcid.org" in a["href"]:
            orcid_url = a["href"]
            if not orcid_url.startswith("http"):
                orcid_url = "https://" + orcid_url.lstrip("/")
            padre = a.find_parent()
            texto_cercano = padre.get_text(strip=True) if padre else ""
            mejor_match = None
            for autor in autores_lista:
                apellido = autor.split()[-1] if autor.split() else ""
                if apellido and (apellido.lower() in texto_cercano.lower() or
                                 apellido.lower() in a.get_text(strip=True).lower()):
                    mejor_match = autor
                    break
            if mejor_match and mejor_match not in orcid_map:
                orcid_map[mejor_match] = orcid_url
            elif not mejor_match:
                orcid_id = orcid_pattern.search(orcid_url)
                if orcid_id and orcid_url not in orcid_map.values():
                    key = f"ORCID_{orcid_id.group(1)}"
                    orcid_map[key] = orcid_url
    return orcid_map


def extraer_afiliaciones_pais(soup, meta_tags, autores_lista):
    afiliaciones = meta_tags.get("afiliaciones_meta", [])
    resultados = []
    if not afiliaciones:
        for cls in ["author-affiliation", "affiliation", "aff", "author-institution"]:
            tags = soup.find_all(class_=re.compile(cls, re.I))
            for tag in tags:
                texto = tag.get_text(strip=True)
                if texto and texto not in afiliaciones:
                    afiliaciones.append(texto)
    for i, aff in enumerate(afiliaciones):
        pais, lat, lon = geocodificar_pais(aff)
        autor_nombre = autores_lista[i] if i < len(autores_lista) else f"Autor {i+1}"
        resultados.append({"autor": autor_nombre, "afiliacion": aff,
                           "pais": pais or "", "lat": lat, "lon": lon})
    return resultados


def scrape_articulo_completo(href):
    try:
        resp = requests.get(href, headers=HEADERS, verify=False, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
    except Exception:
        return {}

    meta_tags = extraer_metadatos_metatags(soup)
    autores_detalle = extraer_autores_con_afiliacion(soup)

    if autores_detalle:
        autores_nombres = [a["nombre"] for a in autores_detalle]
    elif meta_tags["autores_meta"]:
        autores_nombres = meta_tags["autores_meta"]
    else:
        autores_nombres = []
        for cls in ["authors", "author", "article-summary-authors"]:
            tag = soup.find(class_=re.compile(cls, re.I))
            if tag:
                autores_nombres = [a.strip() for a in tag.get_text().split(",") if a.strip()]
                break

    autores_str = ", ".join(autores_nombres) if autores_nombres else ""
    doi = extraer_doi_mejorado(soup, meta_tags, href)
    abstract = extraer_abstract_mejorado(soup, meta_tags)
    keywords = extraer_keywords_mejorado(soup, meta_tags)
    orcid_map = extraer_orcids_mejorado(soup, meta_tags, autores_nombres)

    afiliaciones_detalle = []
    if autores_detalle:
        afiliaciones_detalle = [
            {"autor": a["nombre"], "afiliacion": a["afiliacion"],
             "pais": a["pais"], "lat": a["lat"], "lon": a["lon"]}
            for a in autores_detalle
        ]
    else:
        afiliaciones_detalle = extraer_afiliaciones_pais(soup, meta_tags, autores_nombres)

    paginacion = ""
    if meta_tags["pagina_inicio"] and meta_tags["pagina_fin"]:
        paginacion = f"pp. {meta_tags['pagina_inicio']}–{meta_tags['pagina_fin']}"

    cc = ""
    for a_tag in soup.find_all("a", href=True):
        if "creativecommons.org" in a_tag["href"]:
            cc = a_tag["href"]
            break
    if not cc:
        m_cc = re.search(r"CC[\s-]+BY(?:[-\s][A-Z]+)*", soup.get_text())
        if m_cc:
            cc = m_cc.group()

    copyright_txt = ""
    for cls in ["copyright", "article-copyright"]:
        tag = soup.find(class_=re.compile(cls, re.I))
        if tag:
            copyright_txt = tag.get_text(strip=True)[:300]
            break

    anio_pub = ""
    if meta_tags["fecha_pub"]:
        m_y = re.search(r"\b(19|20)\d{2}\b", meta_tags["fecha_pub"])
        if m_y:
            anio_pub = m_y.group()

    paises_autores = [d["pais"] for d in afiliaciones_detalle if d.get("pais")]
    afiliaciones_str = " | ".join([d["afiliacion"] for d in afiliaciones_detalle if d.get("afiliacion")])

    return {
        "autores": autores_str, "autores_detalle": autores_detalle,
        "afiliaciones": afiliaciones_str, "afiliaciones_detalle": afiliaciones_detalle,
        "paises_autores": paises_autores, "doi": doi,
        "resumen": abstract, "keywords": keywords, "orcids": orcid_map,
        "paginacion": paginacion, "creative_commons": cc,
        "copyright": copyright_txt, "anio_pub": anio_pub,
        "volumen_art": meta_tags.get("volumen_meta", ""),
        "numero_art": meta_tags.get("numero_meta", ""),
        "fecha_pub": meta_tags.get("fecha_pub", ""),
        "journal_meta": meta_tags.get("journal_meta", ""),
        "tiene_doi": bool(doi), "tiene_orcid": bool(orcid_map),
        "tiene_keywords": bool(keywords),
        "tiene_abstract": bool(abstract and len(abstract) > 30),
        "tiene_afiliacion": bool(afiliaciones_str),
        "tiene_pais": bool(paises_autores),
        "tiene_cc": bool(cc), "tiene_paginacion": bool(paginacion),
        "tiene_copyright": bool(copyright_txt),
    }


def extraer_articulos_issue_completo(issue_url, issue_meta=None):
    try:
        resp = requests.get(issue_url, headers=HEADERS, verify=False, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
    except Exception:
        return []

    base_url = "/".join(issue_url.split("/")[:3])
    urls_vistas = set()
    enlaces = []

    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/article/view/" not in href:
            continue
        texto = a.get_text(strip=True)
        if not texto or len(texto) < 5:
            continue
        if href.startswith("/"):
            href = base_url + href
        if re.search(r"/article/view/\d+/\d+$", href):
            continue
        if href not in urls_vistas:
            urls_vistas.add(href)
            padre = (
                a.find_parent(class_=re.compile(r"article.?summary|obj_article", re.I)) or
                a.find_parent("li") or a.find_parent("div")
            )
            autores_preview = ""
            if padre:
                for cls in ["authors", "author", "article-summary-authors"]:
                    at = padre.find(class_=re.compile(cls, re.I))
                    if at:
                        autores_preview = at.get_text(strip=True)
                        break
            enlaces.append({"titulo": texto, "href": href, "autores_preview": autores_preview})

    articulos = []
    for enlace in enlaces:
        meta = scrape_articulo_completo(enlace["href"])
        art = {
            "articulo": enlace["titulo"],
            "autores": meta.get("autores") or enlace["autores_preview"],
            "articulo_url": enlace["href"], **meta,
        }
        if issue_meta:
            art.update({
                "issue": issue_meta.get("issue", ""),
                "anio_issue": issue_meta.get("anio", ""),
                "volumen_issue": issue_meta.get("volumen", ""),
                "numero_issue": issue_meta.get("numero", ""),
                "tipo_issue": issue_meta.get("tipo", ""),
            })
        if not art.get("anio_issue") and art.get("anio_pub"):
            art["anio_issue"] = art["anio_pub"]
        articulos.append(art)
        time.sleep(0.25)
    return articulos


# ═══════════════════════════════════════════════════════
# ── FILTRADO POR PERÍODO ──────────────────────────────
# ═══════════════════════════════════════════════════════

def obtener_anios_disponibles(df):
    col = "anio_issue" if "anio_issue" in df.columns else "anio_pub"
    anios = [a for a in df[col].dropna().unique() if str(a).strip().isdigit() and len(str(a)) == 4]
    return sorted(set(anios))


def filtrar_por_periodo(df, anio_inicio, anio_fin):
    col = "anio_issue" if "anio_issue" in df.columns else "anio_pub"
    df_f = df.copy()
    df_f["_anio_num"] = pd.to_numeric(df_f[col], errors="coerce")
    mask = (df_f["_anio_num"] >= int(anio_inicio)) & (df_f["_anio_num"] <= int(anio_fin))
    return df_f[mask].drop(columns=["_anio_num"]).reset_index(drop=True)


def widget_selector_periodo(df, key_prefix="main"):
    anios = obtener_anios_disponibles(df)
    if not anios:
        return df, None, None

    anio_min = int(min(anios))
    anio_max = int(max(anios))

    st.markdown(f"""
    <div class="period-filter">
        <h4> Filtro de período</h4>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        if anio_min == anio_max:
            st.info(f"Solo hay datos del año {anio_min}")
            rango = (anio_min, anio_max)
        else:
            rango = st.slider(
                "Seleccionar rango de años",
                min_value=anio_min, max_value=anio_max,
                value=(anio_min, anio_max),
                key=f"{key_prefix}_slider_anio",
            )
    with col2:
        st.metric("Año inicio", rango[0])
    with col3:
        st.metric("Año fin", rango[1])

    df_filtrado = filtrar_por_periodo(df, rango[0], rango[1])
    n_total = len(df)
    n_filtrado = len(df_filtrado)
    pct = round(n_filtrado / n_total * 100, 1) if n_total > 0 else 0

    if n_filtrado < n_total:
        st.info(f"📊 **{n_filtrado}** de **{n_total}** artículos ({pct}%) · Período: {rango[0]}–{rango[1]}")

    return df_filtrado, rango[0], rango[1]


# ═══════════════════════════════════════════════════════
# ── GRÁFICAS CON PALETA AZUL ─────────────────────────
# ═══════════════════════════════════════════════════════

def fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    return buf.read()


def generar_nube_autores_bytes(df_aut, max_words=50):
    """Genera nube de palabras optimizada y retorna bytes."""
    if df_aut.empty:
        return None
    try:
        df_top = df_aut.head(max_words)
        frecuencias = {row['autor']: row['num_articulos'] for _, row in df_top.iterrows()}
        wc = WordCloud(
            width=800, height=400,
            background_color=PALETTE["50"],
            colormap="Blues",
            max_words=max_words
        ).generate_from_frequencies(frecuencias)

        fig, ax = plt.subplots(figsize=(8, 4))
        ax.imshow(wc, interpolation="bilinear")
        ax.axis("off")
        fig.patch.set_facecolor(PALETTE["50"])
        plt.tight_layout()
        img_bytes = fig_to_bytes(fig)
        plt.close(fig) # Prevenir memory leaks
        return img_bytes
    except Exception:
        return None


def generar_analisis_temporal(df_art):
    figs = {}
    col_anio = "anio_issue" if "anio_issue" in df_art.columns else "anio_pub"
    df_anio = df_art[df_art[col_anio].astype(str).str.match(r"\d{4}")].copy()

    if not df_anio.empty:
        conteo_anio = (
            df_anio.groupby(col_anio)
            .agg(
                articulos=("articulo", "count"),
                con_doi=("tiene_doi", lambda x: x.fillna(False).astype(bool).sum()),
                con_orcid=("tiene_orcid", lambda x: x.fillna(False).astype(bool).sum()),
                con_abstract=("tiene_abstract", lambda x: x.fillna(False).astype(bool).sum()),
                con_keywords=("tiene_keywords", lambda x: x.fillna(False).astype(bool).sum()),
                autores_unicos=("autores", lambda x: len(set(
                    a.strip() for autores in x.dropna()
                    for a in autores.split(",") if a.strip()
                )))
            )
            .reset_index()
            .sort_values(col_anio)
        )
        figs["conteo_anio"] = conteo_anio

        fig_anio = go.Figure()
        fig_anio.add_trace(go.Bar(
            x=conteo_anio[col_anio].astype(str),
            y=conteo_anio["articulos"],
            marker=dict(
                color=conteo_anio["articulos"],
                colorscale=BLUE_SCALE,
                showscale=True,
                colorbar=dict(title="Artículos", thickness=14),
                line=dict(width=0),
            ),
            text=conteo_anio["articulos"],
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>Artículos: %{y}<extra></extra>",
        ))
        fig_anio.update_layout(
            title=dict(text="Artículos publicados por año", font=dict(size=16, color="#0a0a0a")),
            xaxis=dict(title="Año", tickangle=-45, type="category"),
            yaxis=dict(title="Número de artículos"),
            plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
            margin=dict(t=60, b=60, l=40, r=40),
            font=dict(color="#0a0a0a"),
        )
        figs["fig_anio"] = fig_anio

        fig_meta_anio = go.Figure()
        trazas = [
            ("articulos", "Artículos totales", PALETTE["800"], "lines+markers", 3),
            ("con_doi", "Con DOI", PALETTE["600"], "lines+markers", 2),
            ("con_orcid", "Con ORCID", PALETTE["teal"], "lines+markers", 2),
            ("con_abstract", "Con Abstract", PALETTE["300"], "lines+markers", 1.5),
            ("con_keywords", "Con Keywords", PALETTE["400"], "lines+markers", 1.5),
        ]
        for col_k, name, color, mode, width in trazas:
            if col_k in conteo_anio.columns:
                fig_meta_anio.add_trace(go.Scatter(
                    x=conteo_anio[col_anio].astype(str),
                    y=conteo_anio[col_k],
                    name=name, mode=mode,
                    line=dict(color=color, width=width),
                    marker=dict(size=7, color=color),
                ))
        fig_meta_anio.update_layout(
            title=dict(text="Evolución de metadatos por año", font=dict(size=16, color="#0a0a0a")),
            xaxis=dict(title="Año", type="category"),
            yaxis=dict(title="Artículos"),
            plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
            legend=dict(orientation="h", yanchor="bottom", y=-0.35),
            hovermode="x unified",
            font=dict(color="#0a0a0a"),
        )
        figs["fig_meta_anio"] = fig_meta_anio

        df_pct = conteo_anio.copy()
        df_pct["pct_doi"] = (df_pct["con_doi"] / df_pct["articulos"] * 100).round(1)
        df_pct["pct_orcid"] = (df_pct["con_orcid"] / df_pct["articulos"] * 100).round(1)

        fig_pct = go.Figure()
        fig_pct.add_trace(go.Bar(
            x=df_pct[col_anio].astype(str), y=df_pct["pct_doi"],
            name="% con DOI", marker_color=PALETTE["600"],
            text=df_pct["pct_doi"].apply(lambda x: f"{x}%"),
            textposition="auto",
        ))
        fig_pct.add_trace(go.Bar(
            x=df_pct[col_anio].astype(str), y=df_pct["pct_orcid"],
            name="% con ORCID", marker_color=PALETTE["teal"],
            text=df_pct["pct_orcid"].apply(lambda x: f"{x}%"),
            textposition="auto",
        ))
        fig_pct.update_layout(
            barmode="group",
            title=dict(text="Porcentaje de cobertura DOI y ORCID por año", font=dict(size=16, color="#0a0a0a")),
            xaxis=dict(title="Año", type="category"),
            yaxis=dict(title="% Artículos", range=[0, 110]),
            plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
            legend=dict(orientation="h", yanchor="bottom", y=-0.3),
            font=dict(color="#0a0a0a"),
        )
        figs["fig_pct_anio"] = fig_pct

        campos_heat = ["con_doi", "con_orcid", "con_abstract", "con_keywords"]
        labels_heat = ["DOI", "ORCID", "Abstract", "Keywords"]
        z_vals = []
        for col_k in campos_heat:
            if col_k in conteo_anio.columns:
                pct_col = (conteo_anio[col_k] / conteo_anio["articulos"] * 100).round(1).tolist()
                z_vals.append(pct_col)

        if z_vals:
            fig_heat = go.Figure(data=go.Heatmap(
                z=z_vals,
                x=conteo_anio[col_anio].astype(str).tolist(),
                y=labels_heat[:len(z_vals)],
                colorscale=BLUE_SCALE,
                text=[[f"{v}%" for v in row] for row in z_vals],
                texttemplate="%{text}",
                colorbar=dict(title="% Cobertura"),
            ))
            fig_heat.update_layout(
                title=dict(text="Mapa de calor: cobertura de metadatos por año", font=dict(size=15, color="#0a0a0a")),
                xaxis=dict(title="Año"),
                yaxis=dict(title="Campo"),
                paper_bgcolor="white",
                font=dict(color="#0a0a0a"),
            )
            figs["fig_heatmap"] = fig_heat

    if "volumen_issue" in df_art.columns:
        df_vol = df_art[df_art["volumen_issue"].astype(str).str.strip().ne("")]
        if not df_vol.empty:
            conteo_vol = (
                df_vol.groupby("volumen_issue")
                .agg(articulos=("articulo", "count"))
                .reset_index().sort_values("volumen_issue")
            )
            figs["conteo_vol"] = conteo_vol
            fig_vol = go.Figure()
            fig_vol.add_trace(go.Bar(
                x=conteo_vol["volumen_issue"].astype(str),
                y=conteo_vol["articulos"],
                marker=dict(color=conteo_vol["articulos"], colorscale=BLUE_SCALE),
                text=conteo_vol["articulos"], textposition="outside",
            ))
            fig_vol.update_layout(
                title=dict(text="Artículos por volumen", font=dict(size=15, color="#0a0a0a")),
                xaxis=dict(title="Volumen", type="category"),
                yaxis=dict(title="Artículos"),
                plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
                font=dict(color="#0a0a0a"),
            )
            figs["fig_vol"] = fig_vol

    return figs


def grafica_metadatos_presencia(df, titulo_extra=""):
    campos = {
        "DOI": "tiene_doi", "ORCID": "tiene_orcid", "Keywords": "tiene_keywords",
        "Abstract": "tiene_abstract", "Afiliación": "tiene_afiliacion",
        "País autor": "tiene_pais", "Copyright": "tiene_copyright",
        "Creative Commons": "tiene_cc", "Paginación": "tiene_paginacion",
    }
    nombres, porcentajes, counts = [], [], []
    n = len(df)
    for nombre, col in campos.items():
        if col in df.columns:
            cnt = df[col].fillna(False).astype(bool).sum()
            pct = round(cnt / n * 100, 1) if n > 0 else 0
            nombres.append(nombre)
            porcentajes.append(pct)
            counts.append(int(cnt))

    colores_barra = [
        PALETTE["600"] if p >= 70 else PALETTE["300"] if p >= 30 else PALETTE["100"]
        for p in porcentajes
    ]
    fig = go.Figure(go.Bar(
        x=porcentajes, y=nombres, orientation="h",
        marker=dict(color=colores_barra, line=dict(width=0)),
        text=[f"{p}%  ({c}/{n})" for p, c in zip(porcentajes, counts)],
        textposition="outside",
    ))
    fig.add_vline(x=50, line_dash="dash", line_color="gray", opacity=0.5)
    fig.update_layout(
        title=dict(
            text=f"Presencia de metadatos{' · ' + titulo_extra if titulo_extra else ''}",
            font=dict(size=15, color="#0a0a0a")
        ),
        xaxis=dict(title="% de artículos", range=[0, 130]),
        yaxis=dict(title=""),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        height=380, margin=dict(t=60, b=40, l=120, r=120),
        font=dict(color="#0a0a0a"),
    )
    return fig


def grafica_metadatos_matplotlib(df, titulo_extra=""):
    campos = {
        "DOI": "tiene_doi", "ORCID": "tiene_orcid", "Keywords": "tiene_keywords",
        "Abstract": "tiene_abstract", "Afiliación": "tiene_afiliacion",
        "País autor": "tiene_pais", "Copyright": "tiene_copyright",
        "Creative Commons": "tiene_cc", "Paginación": "tiene_paginacion",
    }
    nombres, porcentajes, counts = [], [], []
    n = len(df)
    for nombre, col in campos.items():
        if col in df.columns:
            cnt = df[col].fillna(False).astype(bool).sum()
            pct = round(cnt / n * 100, 1) if n > 0 else 0
            nombres.append(nombre)
            porcentajes.append(pct)
            counts.append(int(cnt))
    colores = [PALETTE["600"] if p >= 70 else PALETTE["300"] if p >= 30 else PALETTE["100"]
               for p in porcentajes]
    fig, ax = plt.subplots(figsize=(10, 5.5))
    bars = ax.barh(nombres, porcentajes, color=colores, edgecolor="white", linewidth=0.5)
    ax.set_xlim(0, 130)
    ax.set_xlabel("Porcentaje de artículos (%)", fontsize=10)
    titulo = f"Presencia de metadatos{' · ' + titulo_extra if titulo_extra else ''}"
    ax.set_title(titulo, fontsize=12, fontweight="bold", color="#0a0a0a")
    ax.axvline(x=50, color="gray", linestyle="--", alpha=0.4, linewidth=1)
    for bar, pct, cnt in zip(bars, porcentajes, counts):
        ax.text(bar.get_width() + 1, bar.get_y() + bar.get_height()/2,
                f"{pct}%  ({cnt}/{n})", va="center", fontsize=9, color="#374151")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", alpha=0.2)
    plt.tight_layout()
    return fig


def generar_mapa_autores(df_art):
    registros = []
    for _, row in df_art.iterrows():
        afils = row.get("afiliaciones_detalle", [])
        if isinstance(afils, list):
            for aff in afils:
                if aff.get("pais") and aff.get("lat") is not None:
                    registros.append({
                        "autor": aff.get("autor", ""),
                        "pais": aff.get("pais", ""),
                        "afiliacion": aff.get("afiliacion", ""),
                        "lat": aff.get("lat"),
                        "lon": aff.get("lon"),
                        "articulo": row.get("articulo", "")[:60],
                        "anio": row.get("anio_issue", ""),
                    })
    if not registros:
        return None, pd.DataFrame()

    df_mapa = pd.DataFrame(registros)
    df_pais = (
        df_mapa.groupby(["pais", "lat", "lon"])
        .agg(autores=("autor", "count"),
             articulos=("articulo", "nunique"),
             afiliaciones=("afiliacion", lambda x: "<br>".join(x.unique()[:3])))
        .reset_index()
    )

    fig = go.Figure()
    fig.add_trace(go.Scattergeo(
        lat=df_pais["lat"], lon=df_pais["lon"],
        mode="markers+text",
        marker=dict(
            size=df_pais["autores"] * 8 + 10,
            color=df_pais["autores"],
            colorscale=BLUE_SCALE,
            showscale=True,
            colorbar=dict(title="Autores", thickness=15),
            line=dict(width=1, color="white"),
            sizemode="area",
            sizeref=max(df_pais["autores"].max() / 1000, 0.1),
        ),
        text=df_pais["pais"],
        textposition="top center",
        hovertemplate=(
            "<b>%{text}</b><br>Autores: %{marker.color}<br>"
            "Artículos: %{customdata[0]}<extra></extra>"
        ),
        customdata=df_pais[["articulos", "afiliaciones"]].values,
        name="Países",
    ))

    if len(df_pais) > 1:
        conn_pairs = set()
        for _, row in df_art.iterrows():
            afils = row.get("afiliaciones_detalle", [])
            if isinstance(afils, list):
                paises_art = [
                    (a["pais"], a["lat"], a["lon"])
                    for a in afils
                    if a.get("pais") and a.get("lat") is not None
                ]
                for i in range(len(paises_art)):
                    for j in range(i + 1, len(paises_art)):
                        p1, p2 = paises_art[i], paises_art[j]
                        if p1[0] != p2[0]:
                            key = tuple(sorted([p1[0], p2[0]]))
                            conn_pairs.add((key, p1[1], p1[2], p2[1], p2[2]))
        for pair_data in list(conn_pairs)[:50]:
            _, lat1, lon1, lat2, lon2 = pair_data
            fig.add_trace(go.Scattergeo(
                lat=[lat1, lat2, None], lon=[lon1, lon2, None],
                mode="lines",
                line=dict(width=1, color=palette_color("300", 0.4)),
                showlegend=False, hoverinfo="skip",
            ))

    fig.update_layout(
        title=dict(
            text="Distribución geográfica de autores y colaboraciones internacionales",
            x=0.5, font=dict(size=15, color="#0a0a0a"),
        ),
        geo=dict(
            showframe=False, showcoastlines=True, coastlinecolor="lightgray",
            showland=True, landcolor=PALETTE["50"],
            showocean=True, oceancolor="#e8f4fd",
            showcountries=True, countrycolor="white",
            showlakes=True, lakecolor="#e8f4fd",
            projection_type="natural earth",
        ),
        height=550, margin=dict(l=0, r=0, t=50, b=10),
        paper_bgcolor="white",
        font=dict(color="#0a0a0a"),
    )
    return fig, df_mapa


def generar_grafica_paises(df_art):
    registros = []
    for _, row in df_art.iterrows():
        afils = row.get("afiliaciones_detalle", [])
        if isinstance(afils, list):
            for aff in afils:
                if aff.get("pais"):
                    registros.append(aff["pais"])
    if not registros:
        return None
    conteo = Counter(registros)
    df_c = (pd.DataFrame(conteo.items(), columns=["País", "Autores"])
            .sort_values("Autores", ascending=True).tail(20))
    fig = go.Figure(go.Bar(
        x=df_c["Autores"], y=df_c["País"], orientation="h",
        marker=dict(color=df_c["Autores"], colorscale=BLUE_SCALE, line=dict(width=0)),
        text=df_c["Autores"], textposition="outside",
    ))
    fig.update_layout(
        title=dict(text="Top países por número de autores", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="Autores"),
        yaxis=dict(title=""),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        height=max(400, len(df_c) * 28),
        margin=dict(l=140, r=80),
        font=dict(color="#0a0a0a"),
    )
    return fig


# ═══════════════════════════════════════════════════════
# ── ANÁLISIS PRINCIPAL ────────────────────────────────
# ═══════════════════════════════════════════════════════

def limpiar_texto(texto, stop_words):
    if pd.isna(texto) or not texto:
        return ""
    texto = str(texto).lower()
    texto = re.sub(r"http\S+", "", texto)
    texto = re.sub(r"\d+", "", texto)
    texto = re.sub(r"[^\w\s]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    palabras = [w for w in texto.split()
                if w not in stop_words and len(w) > 2 and w not in STOPWORDS_ACADEMICAS]
    return " ".join(palabras)


def detectar_topico_regional(terminos_topico):
    """Detecta si un tópico tiene orientación regional/geográfica."""
    palabras_regionales = {
        "mexico", "méxico", "latinoamerica", "latinoamérica", "america", "argentina",
        "colombia", "chile", "brasil", "peru", "región", "regional", "local",
        "nacional", "municipal", "territorial", "urban", "rural", "ciudad",
        "estado", "provincia", "geografía", "geográfico", "territorio", "zona",
        "espacio", "lugar", "comunidad", "comunidades", "población",
    }
    terminos_lower = {t.lower() for t in terminos_topico}
    coincidencias = terminos_lower.intersection(palabras_regionales)
    return len(coincidencias) >= 2, list(coincidencias)


def analizar_datos(df_art, n_clusters, n_topics, stop_words, progress_cb=None):
    resultados = {}

    if progress_cb:
        progress_cb(0.05, "Estadísticas por número…")
    conteo_issues = df_art.groupby("issue").size().reset_index(name="num_articulos")
    conteo_issues = conteo_issues.sort_values("num_articulos", ascending=False)
    resultados["conteo_issues"] = conteo_issues

    fig1, ax1 = plt.subplots(figsize=(14, 4))
    n_bars = len(conteo_issues)
    bars = ax1.bar(range(n_bars), conteo_issues["num_articulos"],
                   color=PALETTE["600"], edgecolor="white")
    for bar in bars:
        h = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2, h + 0.1, str(int(h)),
                 ha="center", va="bottom", fontsize=7.5)
    ax1.set_xticks(range(n_bars))
    ax1.set_xticklabels(
        [t[:30] + "…" if len(t) > 30 else t for t in conteo_issues["issue"]],
        rotation=45, ha="right", fontsize=6.5
    )
    ax1.set_ylabel("Artículos")
    ax1.set_title("Artículos por número (issue)", fontweight="bold", color="#0a0a0a")
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)
    ax1.set_facecolor(PALETTE["50"])
    plt.tight_layout()
    resultados["fig_issues_bytes"] = fig_to_bytes(fig1)
    plt.close(fig1)

    fig_issues_plotly = go.Figure(go.Bar(
        x=conteo_issues["issue"].apply(lambda t: t[:35] + "…" if len(t) > 35 else t),
        y=conteo_issues["num_articulos"],
        marker=dict(color=conteo_issues["num_articulos"], colorscale=BLUE_SCALE,
                    line=dict(width=0)),
        text=conteo_issues["num_articulos"], textposition="outside",
        hovertemplate="<b>%{x}</b><br>Artículos: %{y}<extra></extra>",
    ))
    fig_issues_plotly.update_layout(
        title=dict(text="Artículos por número (issue)", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="", tickangle=-55, type="category"),
        yaxis=dict(title="Artículos"),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        height=420, font=dict(color="#0a0a0a"),
    )
    resultados["fig_issues_plotly"] = fig_issues_plotly

    if progress_cb:
        progress_cb(0.12, "Análisis temporal…")
    figs_temp = generar_analisis_temporal(df_art)
    resultados["figs_temporales"] = figs_temp

    if progress_cb:
        progress_cb(0.20, "Analizando presencia de metadatos…")
    fig_meta_plotly = grafica_metadatos_presencia(df_art)
    fig_meta_mpl = grafica_metadatos_matplotlib(df_art)
    resultados["fig_metadatos_plotly"] = fig_meta_plotly
    resultados["fig_metadatos"] = fig_to_bytes(fig_meta_mpl)
    plt.close(fig_meta_mpl)

    if progress_cb:
        progress_cb(0.28, "Analizando autores…")
    todos_autores = []
    for x in df_art["autores"].dropna():
        todos_autores.extend([a.strip() for a in x.split(",") if a.strip()])
    conteo_autores = Counter(todos_autores)
    df_autores_count = (
        pd.DataFrame(conteo_autores.items(), columns=["autor", "num_articulos"])
        .sort_values("num_articulos", ascending=False)
        .reset_index(drop=True)
    )
    top20 = df_autores_count.head(20)
    resultados["df_autores_count"] = df_autores_count
    resultados["total_autores"] = len(df_autores_count)

    fig_aut_plotly = go.Figure(go.Bar(
        x=top20["num_articulos"],
        y=top20["autor"],
        orientation="h",
        marker=dict(color=top20["num_articulos"], colorscale=BLUE_SCALE, line=dict(width=0)),
        text=top20["num_articulos"], textposition="outside",
    ))
    fig_aut_plotly.update_layout(
        title=dict(text="Top 20 autores por producción", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="Artículos"),
        yaxis=dict(title="", autorange="reversed"),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        height=500, margin=dict(l=200, r=80),
        font=dict(color="#0a0a0a"),
    )
    resultados["fig_autores_plotly"] = fig_aut_plotly

    fig2, ax2 = plt.subplots(figsize=(10, 5))
    bars2 = ax2.barh(top20["autor"], top20["num_articulos"],
                     color=PALETTE["600"], edgecolor="white")
    for bar in bars2:
        w = bar.get_width()
        ax2.text(w + 0.05, bar.get_y() + bar.get_height()/2, str(int(w)),
                 ha="left", va="center", fontsize=8)
    ax2.invert_yaxis()
    ax2.set_xlabel("Artículos")
    ax2.set_title("Top 20 autores", fontweight="bold", color="#0a0a0a")
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_visible(False)
    ax2.set_facecolor(PALETTE["50"])
    plt.tight_layout()
    resultados["fig_autores"] = fig_to_bytes(fig2)
    plt.close(fig2)

    if progress_cb:
        progress_cb(0.36, "Generando mapa geográfico…")
    fig_mapa, df_mapa = generar_mapa_autores(df_art)
    resultados["fig_mapa"] = fig_mapa
    resultados["df_mapa"] = df_mapa
    fig_paises = generar_grafica_paises(df_art)
    resultados["fig_paises"] = fig_paises

    if progress_cb:
        progress_cb(0.44, "Construyendo matriz TF-IDF…")
    df_art["texto_combinado"] = df_art.apply(
        lambda x: f"{x['articulo']} {x.get('resumen', '')} {x.get('keywords', '')}", axis=1)
    df_art["texto_limpio"] = df_art["texto_combinado"].apply(
        lambda t: limpiar_texto(t, stop_words))

    df_valido = df_art[df_art["texto_limpio"].str.len() > 5].copy()
    if len(df_valido) < 2:
        df_valido = df_art.copy()
        df_valido["texto_limpio"] = df_valido["articulo"].apply(
            lambda t: limpiar_texto(t, stop_words))

    vectorizer = TfidfVectorizer(min_df=1, max_features=3000)
    tfidf_matrix = vectorizer.fit_transform(df_valido["texto_limpio"])
    resultados["vectorizer"] = vectorizer

    if progress_cb:
        progress_cb(0.54, f"KMeans (k={n_clusters})…")
    n_cl = min(n_clusters, len(df_valido))
    kmeans = KMeans(n_clusters=n_cl, random_state=42, n_init=10)
    kmeans.fit(tfidf_matrix)
    df_valido = df_valido.copy()
    df_valido["cluster"] = kmeans.labels_

    terms = vectorizer.get_feature_names_out()
    order_centroids = kmeans.cluster_centers_.argsort()[:, ::-1]
    top_terms_clusters = {i: [terms[ind] for ind in order_centroids[i, :12]]
                          for i in range(n_cl)}
    conteo_clusters = df_valido.groupby("cluster").size().reset_index(name="num_articulos")
    resultados["conteo_clusters"] = conteo_clusters
    resultados["top_terms_clusters"] = top_terms_clusters

    fig_cl_plotly = go.Figure(go.Bar(
        x=[str(c) for c in conteo_clusters["cluster"]],
        y=conteo_clusters["num_articulos"],
        marker=dict(color=conteo_clusters["num_articulos"], colorscale=BLUE_SCALE),
        text=conteo_clusters["num_articulos"], textposition="outside",
    ))
    fig_cl_plotly.update_layout(
        title=dict(text="Distribución por cluster temático", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="Cluster", type="category"),
        yaxis=dict(title="Artículos"),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        font=dict(color="#0a0a0a"),
    )
    resultados["fig_clusters_plotly"] = fig_cl_plotly

    fig3, ax3 = plt.subplots(figsize=(7, 4))
    ax3.bar([str(c) for c in conteo_clusters["cluster"]], conteo_clusters["num_articulos"],
            color=PALETTE["600"], edgecolor="white")
    for i, v in enumerate(conteo_clusters["num_articulos"]):
        ax3.text(i, v + 0.1, str(v), ha="center", va="bottom", fontsize=9)
    ax3.set_xlabel("Cluster")
    ax3.set_ylabel("Artículos")
    ax3.set_title("Distribución por cluster", color="#0a0a0a")
    ax3.spines["top"].set_visible(False)
    ax3.set_facecolor(PALETTE["50"])
    plt.tight_layout()
    resultados["fig_clusters"] = fig_to_bytes(fig3)
    plt.close(fig3)

    if progress_cb:
        progress_cb(0.62, "Proyección PCA…")
    pca = PCA(n_components=2, random_state=42)
    tfidf_pca = pca.fit_transform(tfidf_matrix.toarray())

    cluster_colors = [
        PALETTE["600"], PALETTE["300"], PALETTE["800"], PALETTE["teal"],
        PALETTE["400"], PALETTE["700"], PALETTE["500"], PALETTE["100"],
        PALETTE["900"], PALETTE["200"],
    ]
    fig_pca_plotly = go.Figure()
    for c in range(n_cl):
        mask = df_valido["cluster"] == c
        fig_pca_plotly.add_trace(go.Scatter(
            x=tfidf_pca[mask, 0], y=tfidf_pca[mask, 1],
            mode="markers",
            name=f"Cluster {c}",
            marker=dict(
                size=7, color=cluster_colors[c % len(cluster_colors)],
                line=dict(width=0.5, color="white"), opacity=0.85,
            ),
            text=df_valido[mask]["articulo"].apply(lambda t: t[:50]),
            hovertemplate="<b>Cluster " + str(c) + "</b><br>%{text}<extra></extra>",
        ))
    fig_pca_plotly.update_layout(
        title=dict(text="Proyección PCA de clusters temáticos", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="Componente Principal 1"),
        yaxis=dict(title="Componente Principal 2"),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        font=dict(color="#0a0a0a"),
    )
    resultados["fig_pca_plotly"] = fig_pca_plotly

    fig4, ax4 = plt.subplots(figsize=(9, 6))
    scatter = ax4.scatter(tfidf_pca[:, 0], tfidf_pca[:, 1],
                          c=df_valido["cluster"], cmap="Blues",
                          s=55, alpha=0.8, edgecolors="white", linewidths=0.5)
    plt.colorbar(scatter, ax=ax4, label="Cluster")
    ax4.set_title("PCA de clusters temáticos", color="#0a0a0a")
    ax4.set_xlabel("PC1")
    ax4.set_ylabel("PC2")
    ax4.grid(True, alpha=0.2)
    ax4.set_facecolor(PALETTE["50"])
    plt.tight_layout()
    resultados["fig_pca"] = fig_to_bytes(fig4)
    plt.close(fig4)

    if progress_cb:
        progress_cb(0.70, f"LDA ({n_topics} tópicos)…")
    n_top = min(n_topics, len(df_valido))
    lda = LatentDirichletAllocation(n_components=n_top, random_state=42, max_iter=25)
    lda.fit(tfidf_matrix)
    topic_values = lda.transform(tfidf_matrix)
    df_valido = df_valido.copy()
    df_valido["topico"] = topic_values.argmax(axis=1)
    top_terms_topics = {i: [terms[j] for j in topic.argsort()[:-13:-1]]
                        for i, topic in enumerate(lda.components_)}

    # Detectar tópicos regionales
    topicos_regionales = {}
    for t_idx, t_terms in top_terms_topics.items():
        es_regional, coincidencias = detectar_topico_regional(t_terms)
        if es_regional:
            topicos_regionales[t_idx] = coincidencias

    resultados["top_terms_topics"] = top_terms_topics
    resultados["topicos_regionales"] = topicos_regionales

    conteo_topicos = df_valido.groupby("topico").size().reset_index(name="num_articulos")
    resultados["conteo_topicos"] = conteo_topicos

    wc_figs = []
    for t in range(n_top):
        pesos = lda.components_[t]
        freq_dict = {terms[i]: float(pesos[i]) for i in range(len(terms))}
        try:
            wc = WordCloud(
                width=700, height=350, background_color=PALETTE["50"],
                colormap="Blues", prefer_horizontal=0.9
            ).generate_from_frequencies(freq_dict)
            figw, axw = plt.subplots(figsize=(7, 3.5))
            axw.imshow(wc, interpolation="bilinear")
            axw.axis("off")
            label_regional = " [REGIONAL]" if t in topicos_regionales else ""
            axw.set_title(f"Tópico {t}{label_regional}: {', '.join(top_terms_topics[t][:5])}",
                          fontsize=10, color="#0a0a0a")
            plt.tight_layout()
            wc_figs.append(fig_to_bytes(figw))
            plt.close(figw)
        except Exception:
            wc_figs.append(None)
    resultados["wc_figs"] = wc_figs

    tabla_topicos = pd.crosstab(df_valido["issue"].astype(str), df_valido["topico"])
    resultados["tabla_topicos"] = tabla_topicos

    if progress_cb:
        progress_cb(0.80, "Distribución de tópicos por número…")

    topico_colors = [
        PALETTE["100"], PALETTE["200"], PALETTE["300"], PALETTE["400"], PALETTE["500"],
        PALETTE["600"], PALETTE["700"], PALETTE["800"], PALETTE["900"], PALETTE["teal"],
    ]
    fig_top_plotly = go.Figure()
    for t in range(n_top):
        if t in tabla_topicos.columns:
            name_label = f"Tópico {t}" + (" 🌍" if t in topicos_regionales else "")
            fig_top_plotly.add_trace(go.Bar(
                name=name_label,
                x=tabla_topicos.index.astype(str),
                y=tabla_topicos[t],
                marker_color=topico_colors[t % len(topico_colors)],
            ))
    fig_top_plotly.update_layout(
        barmode="stack",
        title=dict(text="Distribución de tópicos por número (🌍 = regional)", font=dict(size=15, color="#0a0a0a")),
        xaxis=dict(title="Número", tickangle=-45, type="category"),
        yaxis=dict(title="Artículos"),
        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=-0.35),
        height=480,
        font=dict(color="#0a0a0a"),
    )
    resultados["fig_topicos_plotly"] = fig_top_plotly

    fig5, ax5 = plt.subplots(figsize=(14, 5))
    tabla_topicos.plot(kind="bar", stacked=True, ax=ax5,
                       color=[topico_colors[t % len(topico_colors)] for t in range(len(tabla_topicos.columns))])
    ax5.set_title("Distribución de tópicos por número", color="#0a0a0a")
    ax5.set_xlabel("Número")
    ax5.set_ylabel("Artículos")
    plt.xticks(rotation=45, ha="right", fontsize=7)
    ax5.legend(title="Tópico", bbox_to_anchor=(1.05, 1), loc="upper left", fontsize=8)
    ax5.set_facecolor(PALETTE["50"])
    plt.tight_layout()
    resultados["fig_topicos_issue"] = fig_to_bytes(fig5)
    plt.close(fig5)

    if progress_cb:
        progress_cb(0.90, "Red de coautoría…")

    df_valido_copy = df_valido.copy()
    df_valido_copy["autores_lista"] = df_valido_copy["autores"].apply(
        lambda x: [a.strip() for a in str(x).split(",") if a.strip() and len(a.strip()) > 2]
        if pd.notna(x) and x else [])

    edges = []
    for autores in df_valido_copy["autores_lista"]:
        if isinstance(autores, list) and len(autores) > 1:
            edges.extend(list(combinations(autores, 2)))

    # Mapa ORCID global
    orcid_map = {}
    for _, row in df_valido_copy.iterrows():
        orcids_art = row.get("orcids", {})
        if isinstance(orcids_art, dict):
            for autor_key, url in orcids_art.items():
                orcid_map[autor_key] = url
        autores_detalle = row.get("autores_detalle", [])
        if isinstance(autores_detalle, list):
            for adet in autores_detalle:
                if adet.get("orcid") and adet.get("nombre"):
                    orcid_map[adet["nombre"]] = adet["orcid"]

    resultados["orcid_map"] = orcid_map

    G = nx.Graph()
    if edges:
        G.add_edges_from(edges)
        topico_autor = {}
        articulos_autor = {}
        pais_autor = {}
        afil_autor = {}

        for nodo in list(G.nodes()):
            mask = df_valido_copy["autores_lista"].apply(
                lambda lst: nodo in lst if isinstance(lst, list) else False)
            tops = df_valido_copy[mask]["topico"]
            topico_autor[nodo] = int(tops.mode().iloc[0]) if len(tops) > 0 else 0
            articulos_autor[nodo] = int(mask.sum())

            for _, row in df_valido_copy[mask].iterrows():
                afils = row.get("afiliaciones_detalle", [])
                if isinstance(afils, list):
                    for aff in afils:
                        nom = aff.get("autor", "")
                        if nom and any(p.lower() in nodo.lower()
                                       for p in nom.split() if len(p) > 3):
                            if aff.get("pais") and nodo not in pais_autor:
                                pais_autor[nodo] = aff["pais"]
                            if aff.get("afiliacion") and nodo not in afil_autor:
                                afil_autor[nodo] = aff["afiliacion"]
                adet_list = row.get("autores_detalle", [])
                if isinstance(adet_list, list):
                    for adet in adet_list:
                        if adet.get("nombre") and nodo.lower() in adet["nombre"].lower():
                            if adet.get("pais") and nodo not in pais_autor:
                                pais_autor[nodo] = adet["pais"]
                            if adet.get("afiliacion") and nodo not in afil_autor:
                                afil_autor[nodo] = adet["afiliacion"]

        nx.set_node_attributes(G, topico_autor, "topico")
        nx.set_node_attributes(G, articulos_autor, "articulos")
        nx.set_node_attributes(G, orcid_map, "orcid")
        nx.set_node_attributes(G, pais_autor, "pais")
        nx.set_node_attributes(G, afil_autor, "afiliacion")

    resultados["grafo_G"] = G
    resultados["df_articulos"] = df_valido_copy

    resultados["fig_red"] = None
    if G.number_of_edges() > 0:
        topicos_unicos = sorted(df_valido_copy["topico"].dropna().unique())
        blue_shades = [
            PALETTE["100"], PALETTE["300"], PALETTE["500"], PALETTE["600"],
            PALETTE["700"], PALETTE["800"], PALETTE["900"], PALETTE["200"],
            PALETTE["400"], PALETTE["teal"],
        ]
        color_map_dict = {t: blue_shades[i % len(blue_shades)]
                          for i, t in enumerate(topicos_unicos)}
        tamanio_nodo = [max(G.nodes[n].get("articulos", 1) * 90, 40) for n in G.nodes()]
        node_colors = [color_map_dict.get(G.nodes[n].get("topico", 0), blue_shades[0])
                       for n in G.nodes()]

        fig6, ax6 = plt.subplots(figsize=(14, 11))
        pos = nx.spring_layout(G, k=1.0, seed=42)
        nx.draw_networkx_nodes(G, pos, node_color=node_colors,
                               node_size=tamanio_nodo, ax=ax6, alpha=0.9)
        nx.draw_networkx_edges(G, pos, alpha=0.2, ax=ax6, width=0.8,
                               edge_color=PALETTE["300"])
        labels_filtrados = {n: n for n in G.nodes() if G.nodes[n].get("articulos", 1) >= 2}
        nx.draw_networkx_labels(G, pos, labels=labels_filtrados, font_size=6.5, ax=ax6,
                                font_color=PALETTE["900"])
        ax6.set_title("Red de coautoría · tamaño=artículos · color=tópico · 🔗=ORCID: "
                      + str(sum(1 for n in G.nodes() if G.nodes[n].get("orcid"))),
                      fontsize=12, color="#0a0a0a")
        ax6.axis("off")
        ax6.set_facecolor(PALETTE["50"])
        plt.tight_layout()
        resultados["fig_red"] = fig_to_bytes(fig6)
        plt.close(fig6)

    if progress_cb:
        progress_cb(1.0, "Análisis completado ✓")
    return resultados


# ═══════════════════════════════════════════════════════
# ── RED INTERACTIVA CON ORCID ─────────────────────────
# ═══════════════════════════════════════════════════════

def generar_red_interactiva_html(G, orcid_map, df_art):
    if G.number_of_nodes() == 0:
        return "<p style='color:#BCD9F0;padding:2rem;'>No hay datos de coautoría para este período.</p>"

    nodes_data = []
    blue_shades_net = [
        PALETTE["600"], PALETTE["300"], PALETTE["800"], PALETTE["teal"],
        PALETTE["400"], PALETTE["700"], PALETTE["500"], PALETTE["100"],
        PALETTE["900"], PALETTE["200"],
    ]

    for nodo in G.nodes():
        n_arts = G.nodes[nodo].get("articulos", 1)
        topico = G.nodes[nodo].get("topico", 0)
        orcid_url = G.nodes[nodo].get("orcid", "") or orcid_map.get(nodo, "")
        pais = G.nodes[nodo].get("pais", "")
        afil = G.nodes[nodo].get("afiliacion", "")
        color = blue_shades_net[int(topico) % len(blue_shades_net)]
        has_orcid = bool(orcid_url)

        label = nodo[:20] + ("…" if len(nodo) > 20 else "")
        if has_orcid:
            label = "🔗 " + label

        node_obj = {
            "id": nodo, "label": label,
            "value": n_arts * 4,
            "title": (
                f"<div style='font-family:Space Grotesk,sans-serif;padding:8px;min-width:200px;'>"
                f"<b style='color:{PALETTE['600']};font-size:14px;'>{nodo}</b><br>"
                f"<span style='color:#333;'>📄 Artículos: <b>{n_arts}</b></span><br>"
                f"<span style='color:#333;'>🏷️ Tópico: <b>{topico}</b></span><br>"
                f"<span style='color:#333;'>🌍 País: <b>{pais or 'No detectado'}</b></span><br>"
                f"<span style='color:#333;'>🏛️ {afil[:60] + '…' if len(afil) > 60 else afil}</span><br>"
                + (f'<a href="{orcid_url}" target="_blank" style="color:{PALETTE["teal"]};font-weight:600;">'
                   f'🔗 Ver perfil ORCID →</a>'
                   if orcid_url else
                   '<span style="color:#aaa;">Sin ORCID registrado</span>')
                + "</div>"
            ),
            "color": {
                "background": color,
                "border": "#ffffff",
                "highlight": {"background": PALETTE["300"], "border": PALETTE["800"]},
            },
            "font": {"color": "#ffffff", "size": 12, "bold": has_orcid},
            "shadow": has_orcid,
            "orcid": orcid_url, "articulos": n_arts,
            "topico": int(topico), "pais": pais, "afiliacion": afil,
            "has_orcid": has_orcid,
        }
        nodes_data.append(node_obj)

    edges_data = [
        {"from": u, "to": v,
         "color": {"color": f"rgba(4,138,191,0.3)", "hover": f"rgba(4,138,191,0.8)"},
         "width": 1.2}
        for u, v in G.edges()
    ]
    nodes_json = json.dumps(nodes_data, ensure_ascii=False)
    edges_json = json.dumps(edges_data, ensure_ascii=False)

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600&family=JetBrains+Mono&display=swap');
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{background:{PALETTE["800"]};font-family:'Space Grotesk',sans-serif;color:{PALETTE["100"]};}}
#network{{width:100%;height:520px;background:{PALETTE["900"]};border-radius:0 0 12px 12px;}}
#controls{{
  display:flex;gap:8px;align-items:center;flex-wrap:wrap;
  padding:10px 14px;background:{PALETTE["700"]};
  border-radius:12px 12px 0 0;border-bottom:1px solid {PALETTE["500"]};
}}
#controls input{{
  flex:1;min-width:200px;padding:7px 12px;
  background:{PALETTE["800"]};border:1px solid {PALETTE["500"]};
  border-radius:6px;color:{PALETTE["100"]};font-size:13px;
}}
#controls input::placeholder{{color:{PALETTE["300"]};opacity:0.7;}}
.btn{{
  padding:7px 14px;background:{PALETTE["600"]};border:none;border-radius:6px;
  color:white;cursor:pointer;font-size:13px;font-weight:600;
  transition:all 0.2s ease;
}}
.btn:hover{{background:{PALETTE["500"]};box-shadow:0 3px 12px rgba(4,178,217,0.3);}}
.btn.teal{{background:{PALETTE["teal"]};color:white;}}
.btn.teal:hover{{background:{PALETTE["teal_light"]};}}
#stats-bar{{
  display:flex;gap:16px;padding:6px 14px;
  background:{PALETTE["900"]};border-bottom:1px solid {PALETTE["800"]};
  font-family:'JetBrains Mono',monospace;font-size:11px;color:{PALETTE["300"]};flex-wrap:wrap;
}}
.stat-item span{{color:{PALETTE["400"]};font-weight:600;}}
#info-panel{{
  background:{PALETTE["800"]};border-radius:0 0 12px 12px;padding:12px 14px;
  font-size:12px;min-height:60px;border-top:1px solid {PALETTE["700"]};
}}
#info-panel h4{{margin:0 0 6px;color:{PALETTE["300"]};font-size:14px;font-weight:600;}}
.tag{{
  display:inline-block;padding:2px 8px;border-radius:10px;
  background:{PALETTE["700"]};border:1px solid {PALETTE["500"]};
  margin:2px;font-size:11px;color:{PALETTE["100"]};
}}
.orcid-badge{{
  display:inline-block;padding:2px 10px;border-radius:10px;
  background:rgba(3,140,127,0.2);border:1px solid {PALETTE["teal"]};
  margin:2px;font-size:11px;color:{PALETTE["teal_light"]};
}}
</style>
<script src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
</head><body>
<div id="controls">
  <input type="text" id="search-input" placeholder="🔍 Buscar autor, país o institución…" oninput="filtrarAutor()">
  <button class="btn" onclick="resetearVista()">↺ Resetear</button>
  <button class="btn" onclick="mostrarTodos()">⊕ Centrar</button>
  <button class="btn teal" onclick="resaltarConOrcid()">🔗 Solo con ORCID</button>
  <span style="font-size:11px;color:{PALETTE["300"]};margin-left:auto;font-family:'JetBrains Mono',monospace;">
    Clic=detalles · Doble clic=ORCID
  </span>
</div>
<div id="stats-bar">
  <div class="stat-item">Nodos: <span id="s-nodes">0</span></div>
  <div class="stat-item">Aristas: <span id="s-edges">0</span></div>
  <div class="stat-item">Con ORCID: <span id="s-orcid">0</span></div>
  <div class="stat-item">Países: <span id="s-paises">0</span></div>
  <div class="stat-item">Seleccionado: <span id="s-sel">—</span></div>
</div>
<div id="network"></div>
<div id="info-panel">
  <span style="color:{PALETTE["300"]};">↑ Selecciona un nodo para ver detalles y colaboraciones.</span>
</div>
<script>
const nodesData={nodes_json};
const edgesData={edges_json};
const nodes=new vis.DataSet(nodesData);
const edges=new vis.DataSet(edgesData);
const container=document.getElementById('network');
const network=new vis.Network(container,{{nodes,edges}},{{
  nodes:{{
    shape:'dot',
    scaling:{{min:8,max:45,label:{{enabled:true,min:10,max:16}}}},
    borderWidth:2,
    shadow:{{enabled:true,color:'rgba(0,0,0,0.4)',size:8,x:2,y:2}},
  }},
  edges:{{
    smooth:{{type:'continuous',roundness:0.3}},
    width:1.2,hoverWidth:3,selectionWidth:3,
  }},
  physics:{{
    solver:'forceAtlas2Based',
    forceAtlas2Based:{{
      gravitationalConstant:-55,springLength:110,springConstant:0.04,
      damping:0.9,avoidOverlap:0.3,
    }},
    stabilization:{{iterations:250,fit:true}},
  }},
  interaction:{{hover:true,tooltipDelay:80,zoomView:true,dragView:true}},
}});

const totalOrcid=nodesData.filter(n=>n.has_orcid).length;
const paisesSet=new Set(nodesData.map(n=>n.pais).filter(Boolean));
document.getElementById('s-nodes').textContent=nodesData.length;
document.getElementById('s-edges').textContent=edgesData.length;
document.getElementById('s-orcid').textContent=totalOrcid;
document.getElementById('s-paises').textContent=paisesSet.size;

const adjMap={{}};
nodesData.forEach(n=>{{adjMap[n.id]=new Set();}});
edgesData.forEach(e=>{{adjMap[e.from]?.add(e.to);adjMap[e.to]?.add(e.from);}});

network.on('click',function(p){{
  if(p.nodes.length>0){{
    const nodeId=p.nodes[0];
    const ni=nodesData.find(n=>n.id===nodeId);
    if(!ni) return;
    const vecinos=adjMap[nodeId]||new Set();
    const rel=new Set([nodeId,...vecinos]);
    nodes.update(nodesData.map(n=>({{...n,opacity:rel.has(n.id)?1.0:0.12}})));
    const vecList=[...vecinos].slice(0,12).map(v=>
      `<span class="tag">👤 ${{v.length>20?v.substr(0,20)+'…':v}}</span>`
    ).join(" ");
    const orcidHtml=ni.orcid
      ?`<a href="${{ni.orcid}}" target="_blank" class="orcid-badge">🔗 Ver perfil ORCID →</a>`
      :'<span style="color:#aaa;font-size:11px;">Sin ORCID registrado</span>';
    document.getElementById('s-sel').textContent=nodeId.length>20?nodeId.substr(0,20)+'…':nodeId;
    document.getElementById('info-panel').innerHTML=`
      <h4>👤 ${{nodeId}}</h4>
      <div style="margin-bottom:6px;color:{PALETTE["100"]};">
        📄 <b>${{ni.articulos}}</b> artículos · 🏷️ Tópico <b>${{ni.topico}}</b> · 🌍 <b>${{ni.pais||'Sin país'}}</b>
      </div>
      ${{ni.afiliacion?`<div style="color:{PALETTE["200"]};font-size:11px;margin-bottom:6px;">🏛️ ${{ni.afiliacion}}</div>`:''}}
      <div style="margin-bottom:6px;">${{orcidHtml}}</div>
      <div><b style="color:{PALETTE["300"]};">Co-autores (${{vecinos.size}}):</b> ${{vecList||"<span style='color:#aaa;'>Ninguno</span>"}}</div>`;
  }}else{{
    resetearVista();
    document.getElementById('s-sel').textContent='—';
  }}
}});

network.on('doubleClick',function(p){{
  if(p.nodes.length>0){{
    const ni=nodesData.find(n=>n.id===p.nodes[0]);
    if(ni&&ni.orcid)window.open(ni.orcid,'_blank');
    else alert('Este autor no tiene ORCID registrado.');
  }}
}});

function resetearVista(){{
  nodes.update(nodesData.map(n=>({{...n,opacity:1.0}})));
  document.getElementById('info-panel').innerHTML='<span style="color:{PALETTE["300"]};">↑ Selecciona un nodo para ver detalles.</span>';
  network.fit({{animation:{{duration:500,easingFunction:'easeInOutQuad'}}}});
}}
function mostrarTodos(){{network.fit({{animation:{{duration:400}}}});}}
function resaltarConOrcid(){{
  const conOrcid=new Set(nodesData.filter(n=>n.has_orcid).map(n=>n.id));
  if(conOrcid.size===0){{alert('No se encontraron autores con ORCID.');return;}}
  nodes.update(nodesData.map(n=>({{...n,opacity:conOrcid.has(n.id)?1.0:0.08}})));
  document.getElementById('info-panel').innerHTML=
    `<span style="color:{PALETTE["400"]};">🔗 <b>${{conOrcid.size}}</b> autores con ORCID. Doble clic para abrir perfil.</span>`;
}}
function filtrarAutor(){{
  const q=document.getElementById('search-input').value.toLowerCase().trim();
  if(!q){{resetearVista();return;}}
  const coincide=nodesData.filter(n=>
    n.id.toLowerCase().includes(q)||
    (n.pais&&n.pais.toLowerCase().includes(q))||
    (n.afiliacion&&n.afiliacion.toLowerCase().includes(q))
  );
  if(!coincide.length)return;
  const rel=new Set();
  coincide.forEach(n=>{{rel.add(n.id);(adjMap[n.id]||new Set()).forEach(v=>rel.add(v));}});
  nodes.update(nodesData.map(n=>({{...n,opacity:rel.has(n.id)?1.0:0.08}})));
  if(coincide.length===1){{
    network.selectNodes([coincide[0].id]);
    network.focus(coincide[0].id,{{scale:1.5,animation:{{duration:500}}}});
  }}
}}
network.on('stabilizationIterationsDone',function(){{network.fit();}});
</script></body></html>"""
    return html


# ═══════════════════════════════════════════════════════
# ── EXPORTAR EXCEL ────────────────────────────────────
# ═══════════════════════════════════════════════════════

def estilo_header_xlsx(ws, fila, col_inicio, col_fin, titulo=None):
    """Aplica estilo de encabezado a una fila."""
    fill_header = PatternFill("solid", fgColor="055BA6")
    font_header = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alin_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(col_inicio, col_fin + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alin_center

    # Bordes
    borde = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="medium", color="04B2D9"),
    )
    for col in range(col_inicio, col_fin + 1):
        ws.cell(row=fila, column=col).border = borde


def estilo_fila_datos(ws, fila, col_inicio, col_fin, alternado=False):
    """Aplica estilo a filas de datos."""
    fill_alt = PatternFill("solid", fgColor="E8F4FB")
    font_datos = Font(name="Arial", size=9, color="0a0a0a")
    alin = Alignment(vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style="thin", color="BCD9F0"),
        right=Side(style="thin", color="BCD9F0"),
        bottom=Side(style="thin", color="BCD9F0"),
    )
    for col in range(col_inicio, col_fin + 1):
        cell = ws.cell(row=fila, column=col)
        if alternado:
            cell.fill = fill_alt
        cell.font = font_datos
        cell.alignment = alin
        cell.border = borde


def agregar_titulo_hoja(ws, titulo, subtitulo=""):
    """Agrega título vistoso al inicio de la hoja."""
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A1:J1")
    cell_titulo = ws["A1"]
    cell_titulo.value = titulo
    cell_titulo.fill = PatternFill("solid", fgColor="033A5A")
    cell_titulo.font = Font(bold=True, color="04B2D9", name="Arial", size=13)
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    if subtitulo:
        ws.merge_cells("A2:J2")
        cell_sub = ws["A2"]
        cell_sub.value = subtitulo
        cell_sub.fill = PatternFill("solid", fgColor="055BA6")
        cell_sub.font = Font(color="FFFFFF", name="Arial", size=9)
        cell_sub.alignment = Alignment(horizontal="center", vertical="center")


def generar_excel(df_art, resultados, periodo_str=""):
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía default

    # ── Hoja 1: Resumen General ─────────────────────────
    ws1 = wb.create_sheet("Resumen General")
    agregar_titulo_hoja(ws1, "📊 Resumen General del Análisis",
                        f"Período: {periodo_str} · Total artículos: {len(df_art)}")

    campos_meta = {
        "Total de artículos": len(df_art),
        "Números (issues) analizados": df_art["issue"].nunique() if "issue" in df_art.columns else "—",
        "Autores únicos": resultados.get("total_autores", "—"),
        "Con DOI": int(df_art["tiene_doi"].fillna(False).astype(bool).sum()) if "tiene_doi" in df_art.columns else "—",
        "Con ORCID": int(df_art["tiene_orcid"].fillna(False).astype(bool).sum()) if "tiene_orcid" in df_art.columns else "—",
        "Con Abstract": int(df_art["tiene_abstract"].fillna(False).astype(bool).sum()) if "tiene_abstract" in df_art.columns else "—",
        "Con Keywords": int(df_art["tiene_keywords"].fillna(False).astype(bool).sum()) if "tiene_keywords" in df_art.columns else "—",
        "Con Afiliación": int(df_art["tiene_afiliacion"].fillna(False).astype(bool).sum()) if "tiene_afiliacion" in df_art.columns else "—",
        "Con País detectado": int(df_art["tiene_pais"].fillna(False).astype(bool).sum()) if "tiene_pais" in df_art.columns else "—",
        "Autores con ORCID en red": len(resultados.get("orcid_map", {})),
        "Período analizado": periodo_str,
    }

    fila = 4
    estilo_header_xlsx(ws1, fila, 1, 3, "Resumen")
    ws1.cell(row=fila, column=1).value = "Indicador"
    ws1.cell(row=fila, column=2).value = "Valor"
    ws1.cell(row=fila, column=3).value = "Descripción"

    for i, (k, v) in enumerate(campos_meta.items(), fila + 1):
        ws1.cell(row=i, column=1).value = k
        ws1.cell(row=i, column=2).value = v
        ws1.cell(row=i, column=3).value = ""
        estilo_fila_datos(ws1, i, 1, 3, alternado=(i % 2 == 0))
        ws1.cell(row=i, column=1).font = Font(bold=True, name="Arial", size=9, color="033A5A")

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 40

    # ── Hoja 2: Artículos Completos ─────────────────────
    ws2 = wb.create_sheet("Artículos")
    agregar_titulo_hoja(ws2, "📄 Listado Completo de Artículos",
                        f"Exportado: {time.strftime('%d/%m/%Y %H:%M')} · {len(df_art)} artículos")

    cols_art = [
        ("Año", "anio_issue"), ("Vol.", "volumen_issue"), ("Núm.", "numero_issue"),
        ("Número (Issue)", "issue"), ("Título del artículo", "articulo"),
        ("Autores", "autores"), ("DOI", "doi"), ("Keywords", "keywords"),
        ("Resumen", "resumen"), ("Afiliaciones", "afiliaciones"),
        ("Paginación", "paginacion"), ("Creative Commons", "creative_commons"),
        ("Con DOI", "tiene_doi"), ("Con ORCID", "tiene_orcid"),
        ("Con Abstract", "tiene_abstract"), ("Con Keywords", "tiene_keywords"),
        ("URL", "articulo_url"),
    ]
    cols_art = [(label, col) for label, col in cols_art if col in df_art.columns]

    fila = 4
    estilo_header_xlsx(ws2, fila, 1, len(cols_art))
    for j, (label, _) in enumerate(cols_art, 1):
        ws2.cell(row=fila, column=j).value = label

    for i, (_, row) in enumerate(df_art.iterrows(), fila + 1):
        for j, (_, col) in enumerate(cols_art, 1):
            val = row.get(col, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            elif isinstance(val, (list, dict)):
                val = str(val)[:200]
            ws2.cell(row=i, column=j).value = str(val)[:500] if val else ""
        estilo_fila_datos(ws2, i, 1, len(cols_art), alternado=(i % 2 == 0))

    for j, (label, _) in enumerate(cols_art, 1):
        if "Título" in label or "Resumen" in label:
            ws2.column_dimensions[get_column_letter(j)].width = 45
        elif "Autores" in label or "Keywords" in label:
            ws2.column_dimensions[get_column_letter(j)].width = 30
        else:
            ws2.column_dimensions[get_column_letter(j)].width = 15

    ws2.freeze_panes = "A6"

    # ── Hoja 3: Estadísticas por Año ─────────────────────
    ws3 = wb.create_sheet("Por Año")
    agregar_titulo_hoja(ws3, "📅 Estadísticas por Año")

    figs_temp = resultados.get("figs_temporales", {})
    if "conteo_anio" in figs_temp:
        conteo_a = figs_temp["conteo_anio"]
        col_anio_k = [c for c in conteo_a.columns if "anio" in c][0]
        headers_a = ["Año", "Artículos", "Con DOI", "Con ORCID", "Con Abstract", "Con Keywords", "Autores únicos"]
        fila = 4
        estilo_header_xlsx(ws3, fila, 1, len(headers_a))
        for j, h in enumerate(headers_a, 1):
            ws3.cell(row=fila, column=j).value = h

        for i, (_, row) in enumerate(conteo_a.iterrows(), fila + 1):
            row_vals = [
                str(row.get(col_anio_k, "")),
                row.get("articulos", 0),
                row.get("con_doi", 0),
                row.get("con_orcid", 0),
                row.get("con_abstract", 0),
                row.get("con_keywords", 0),
                row.get("autores_unicos", 0),
            ]
            for j, val in enumerate(row_vals, 1):
                ws3.cell(row=i, column=j).value = val
            estilo_fila_datos(ws3, i, 1, len(headers_a), alternado=(i % 2 == 0))

        for j in range(1, len(headers_a) + 1):
            ws3.column_dimensions[get_column_letter(j)].width = 18

    # ── Hoja 4: Estadísticas por Número ──────────────────
    ws4 = wb.create_sheet("Por Número")
    agregar_titulo_hoja(ws4, "Artículos por Número (Issue)")

    conteo_iss = resultados.get("conteo_issues", pd.DataFrame())
    if not conteo_iss.empty:
        headers_iss = ["Número (Issue)", "Artículos publicados"]
        fila = 4
        estilo_header_xlsx(ws4, fila, 1, len(headers_iss))
        for j, h in enumerate(headers_iss, 1):
            ws4.cell(row=fila, column=j).value = h

        for i, (_, row) in enumerate(conteo_iss.iterrows(), fila + 1):
            ws4.cell(row=i, column=1).value = str(row.get("issue", ""))
            ws4.cell(row=i, column=2).value = int(row.get("num_articulos", 0))
            estilo_fila_datos(ws4, i, 1, 2, alternado=(i % 2 == 0))

        ws4.column_dimensions["A"].width = 50
        ws4.column_dimensions["B"].width = 22

    # ── Hoja 5: Autores ───────────────────────────────────
    ws5 = wb.create_sheet("Autores")
    agregar_titulo_hoja(ws5, "👥 Ranking de Autores por Producción")

    df_aut = resultados.get("df_autores_count", pd.DataFrame())
    orcid_map = resultados.get("orcid_map", {})
    if not df_aut.empty:
        headers_aut = ["#", "Autor", "Artículos", "ORCID URL"]
        fila = 4
        estilo_header_xlsx(ws5, fila, 1, len(headers_aut))
        for j, h in enumerate(headers_aut, 1):
            ws5.cell(row=fila, column=j).value = h

        for i, (_, row) in enumerate(df_aut.iterrows(), fila + 1):
            ws5.cell(row=i, column=1).value = i - fila
            ws5.cell(row=i, column=2).value = str(row.get("autor", ""))
            ws5.cell(row=i, column=3).value = int(row.get("num_articulos", 0))
            orcid_url = orcid_map.get(row.get("autor", ""), "")
            ws5.cell(row=i, column=4).value = str(orcid_url) if orcid_url else "—"
            if orcid_url:
                ws5.cell(row=i, column=4).font = Font(
                    color="038C7F", underline="single", name="Arial", size=9)
            estilo_fila_datos(ws5, i, 1, 4, alternado=(i % 2 == 0))

        ws5.column_dimensions["A"].width = 6
        ws5.column_dimensions["B"].width = 35
        ws5.column_dimensions["C"].width = 12
        ws5.column_dimensions["D"].width = 45

    # ── Hoja 6: Distribución Geográfica ──────────────────
    ws6 = wb.create_sheet("Geografía")
    agregar_titulo_hoja(ws6, "Distribución Geográfica de Autores")

    df_mapa = resultados.get("df_mapa", pd.DataFrame())
    if not df_mapa.empty:
        paises_count = Counter(df_mapa["pais"].tolist())
        total_geo = sum(paises_count.values())
        headers_geo = ["País", "Autores", "% del total", "Artículos únicos"]
        fila = 4
        estilo_header_xlsx(ws6, fila, 1, len(headers_geo))
        for j, h in enumerate(headers_geo, 1):
            ws6.cell(row=fila, column=j).value = h

        for i, (pais, cnt) in enumerate(paises_count.most_common(), fila + 1):
            pct = round(cnt / total_geo * 100, 1) if total_geo else 0
            arts_pais = df_mapa[df_mapa["pais"] == pais]["articulo"].nunique()
            vals = [pais, cnt, f"{pct}%", arts_pais]
            for j, val in enumerate(vals, 1):
                ws6.cell(row=i, column=j).value = val
            estilo_fila_datos(ws6, i, 1, len(headers_geo), alternado=(i % 2 == 0))

        for j, w in zip(range(1, 5), [25, 12, 14, 18]):
            ws6.column_dimensions[get_column_letter(j)].width = w

    # ── Hoja 7: Clusters Temáticos ───────────────────────
    ws7 = wb.create_sheet("Clusters")
    agregar_titulo_hoja(ws7, "Clusters Temáticos (KMeans + TF-IDF)",
                        "Agrupación automática de artículos por similitud textual")

    conteo_cl = resultados.get("conteo_clusters", pd.DataFrame())
    top_terms_cl = resultados.get("top_terms_clusters", {})
    if not conteo_cl.empty:
        headers_cl = ["Cluster", "Artículos", "Términos representativos (top 12)"]
        fila = 4
        estilo_header_xlsx(ws7, fila, 1, len(headers_cl))
        for j, h in enumerate(headers_cl, 1):
            ws7.cell(row=fila, column=j).value = h

        for i, (_, row) in enumerate(conteo_cl.iterrows(), fila + 1):
            c = int(row.get("cluster", i - fila - 1))
            terms_str = ", ".join(top_terms_cl.get(c, [])[:12])
            ws7.cell(row=i, column=1).value = f"Cluster {c}"
            ws7.cell(row=i, column=2).value = int(row.get("num_articulos", 0))
            ws7.cell(row=i, column=3).value = terms_str
            estilo_fila_datos(ws7, i, 1, 3, alternado=(i % 2 == 0))

        ws7.column_dimensions["A"].width = 12
        ws7.column_dimensions["B"].width = 12
        ws7.column_dimensions["C"].width = 60

    # ── Hoja 8: Tópicos LDA ───────────────────────────────
    ws8 = wb.create_sheet("Tópicos LDA")
    agregar_titulo_hoja(ws8, "🏷️ Modelado de Tópicos (LDA)",
                        "Distribución probabilística de temas · = tópico regional")

    top_terms_top = resultados.get("top_terms_topics", {})
    conteo_top = resultados.get("conteo_topicos", pd.DataFrame())
    topicos_reg = resultados.get("topicos_regionales", {})

    if top_terms_top:
        headers_top = ["Tópico", "Artículos", "¿Regional?", "Términos principales (top 12)"]
        fila = 4
        estilo_header_xlsx(ws8, fila, 1, len(headers_top))
        for j, h in enumerate(headers_top, 1):
            ws8.cell(row=fila, column=j).value = h

        for t_idx, terms_list in sorted(top_terms_top.items()):
            i = fila + t_idx + 1
            n_arts_top = conteo_top[conteo_top["topico"] == t_idx]["num_articulos"].values
            is_regional = t_idx in topicos_reg
            ws8.cell(row=i, column=1).value = f"Tópico {t_idx}"
            ws8.cell(row=i, column=2).value = int(n_arts_top[0]) if len(n_arts_top) > 0 else 0
            ws8.cell(row=i, column=3).value = "Sí" if is_regional else "No"
            ws8.cell(row=i, column=4).value = ", ".join(terms_list[:12])
            estilo_fila_datos(ws8, i, 1, len(headers_top), alternado=(t_idx % 2 == 0))
            if is_regional:
                ws8.cell(row=i, column=3).font = Font(
                    color="038C7F", bold=True, name="Arial", size=9)

        for j, w in zip(range(1, 5), [12, 12, 14, 60]):
            ws8.column_dimensions[get_column_letter(j)].width = w

    # ── Hoja 9: Red de Coautoría ──────────────────────────
    ws9 = wb.create_sheet("Red de Coautoría")
    agregar_titulo_hoja(ws9, "🕸️ Métricas de Red de Coautoría",
                        "Nodos con ORCID, País y Afiliación detectados automáticamente")

    G = resultados.get("grafo_G", nx.Graph())
    if G.number_of_nodes() > 0:
        headers_red = ["Autor", "Artículos", "Tópico", "Co-autores", "País", "Afiliación", "ORCID URL"]
        fila = 4
        estilo_header_xlsx(ws9, fila, 1, len(headers_red))
        for j, h in enumerate(headers_red, 1):
            ws9.cell(row=fila, column=j).value = h

        nodos_sorted = sorted(G.nodes(), key=lambda n: G.nodes[n].get("articulos", 0), reverse=True)
        for i, nodo in enumerate(nodos_sorted, fila + 1):
            orcid_url = G.nodes[nodo].get("orcid", "") or orcid_map.get(nodo, "")
            vals = [
                nodo,
                G.nodes[nodo].get("articulos", 0),
                G.nodes[nodo].get("topico", "—"),
                G.degree(nodo),
                G.nodes[nodo].get("pais", ""),
                G.nodes[nodo].get("afiliacion", "")[:100],
                orcid_url or "—",
            ]
            for j, val in enumerate(vals, 1):
                ws9.cell(row=i, column=j).value = val
            estilo_fila_datos(ws9, i, 1, len(headers_red), alternado=(i % 2 == 0))
            if orcid_url:
                ws9.cell(row=i, column=7).font = Font(
                    color="038C7F", underline="single", name="Arial", size=9)

        for j, w in zip(range(1, 8), [30, 10, 10, 12, 20, 40, 45]):
            ws9.column_dimensions[get_column_letter(j)].width = w

    # ── Hoja 10: Metadatos por Artículo ──────────────────
    ws10 = wb.create_sheet("Metadatos")
    agregar_titulo_hoja(ws10, "🔖 Cobertura de Metadatos por Artículo")

    campos_bool = [
        ("DOI", "tiene_doi"), ("ORCID", "tiene_orcid"), ("Abstract", "tiene_abstract"),
        ("Keywords", "tiene_keywords"), ("Afiliación", "tiene_afiliacion"),
        ("País", "tiene_pais"), ("Creative Commons", "tiene_cc"),
        ("Paginación", "tiene_paginacion"), ("Copyright", "tiene_copyright"),
    ]
    cols_meta = [("Título", "articulo"), ("Año", "anio_issue"), ("Autores", "autores")] + \
                [(label, col) for label, col in campos_bool if col in df_art.columns]

    fila = 4
    estilo_header_xlsx(ws10, fila, 1, len(cols_meta))
    for j, (label, _) in enumerate(cols_meta, 1):
        ws10.cell(row=fila, column=j).value = label

    for i, (_, row) in enumerate(df_art.iterrows(), fila + 1):
        for j, (_, col) in enumerate(cols_meta, 1):
            val = row.get(col, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            ws10.cell(row=i, column=j).value = str(val)[:200] if val else ""
        estilo_fila_datos(ws10, i, 1, len(cols_meta), alternado=(i % 2 == 0))

    for j in range(1, len(cols_meta) + 1):
        ws10.column_dimensions[get_column_letter(j)].width = 20 if j <= 3 else 14

    # Guardar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════
# ── PDF CON FIGURAS Y EXPLICACIONES ──────────────────
# ═══════════════════════════════════════════════════════

def generar_pdf(url_revista, df_art, resultados, periodo_str=""):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.8*rcm, leftMargin=1.8*rcm,
        topMargin=2*rcm, bottomMargin=2*rcm,
        title="Análisis de Revista OJS v5",
    )
    styles = getSampleStyleSheet()

    c_azul_oscuro  = colors.HexColor("#033A5A")
    c_azul_medio   = colors.HexColor("#055BA6")
    c_azul_claro   = colors.HexColor("#048ABF")
    c_teal         = colors.HexColor("#038C7F")
    c_azul_palido  = colors.HexColor("#E8F4FB")
    c_negro        = colors.HexColor("#0a0a0a")
    c_gris         = colors.HexColor("#374151")
    c_blanco       = colors.white

    title_style = ParagraphStyle("CT", parent=styles["Title"],
        fontSize=22, textColor=c_azul_oscuro, spaceAfter=8,
        alignment=TA_CENTER, fontName="Helvetica-Bold")
    subtitle_style = ParagraphStyle("CS", parent=styles["Normal"],
        fontSize=10, textColor=c_gris,
        alignment=TA_CENTER, spaceAfter=4)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
        fontSize=11, textColor=c_azul_medio, spaceBefore=10,
        spaceAfter=5, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", parent=styles["Normal"],
        fontSize=9, leading=14, alignment=TA_JUSTIFY, textColor=c_negro)
    explain_style = ParagraphStyle("Exp", parent=styles["Normal"],
        fontSize=9, leading=14, textColor=c_negro,
        backColor=colors.HexColor("#E8F4FB"),
        borderPadding=8, leftIndent=8, rightIndent=8)
    center_style = ParagraphStyle("Center", parent=styles["Normal"],
        alignment=TA_CENTER, fontSize=8, textColor=c_gris)
    caption_style = ParagraphStyle("Cap", parent=styles["Normal"],
        fontSize=8.5, leading=12, textColor=c_gris,
        alignment=TA_CENTER, spaceBefore=4, spaceAfter=8,
        fontName="Helvetica-Oblique")

    def section_header(texto, color=None):
        bg = color or c_azul_oscuro
        t = Table(
            [[Paragraph(f"  {texto}", ParagraphStyle("HT",
              fontSize=11, textColor=c_blanco,
              fontName="Helvetica-Bold", leading=16))]],
            colWidths=[17.6*rcm]
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), bg),
            ("TOPPADDING", (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
        ]))
        return t

    story = []

    # ── PORTADA ──────────────────────────────────────────
    story.append(Spacer(1, 0.6*inch))
    story.append(Paragraph("🔵 Análisis de Revista OJS", title_style))
    story.append(Paragraph("Informe de figuras y análisis de contenido · v5.0", subtitle_style))
    story.append(Spacer(1, 0.15*inch))
    story.append(HRFlowable(width="100%", thickness=3, color=c_azul_medio))
    story.append(Spacer(1, 0.2*inch))

    col_anio = "anio_issue" if "anio_issue" in df_art.columns else "anio_pub"
    anios = sorted([a for a in df_art[col_anio].dropna().unique()
                    if str(a).strip().isdigit()]) if col_anio in df_art.columns else []
    anio_str = periodo_str or (f"{anios[0]}–{anios[-1]}" if len(anios) > 1 else (anios[0] if anios else "—"))

    meta_data = [
        ["Revista analizada:", Paragraph(url_revista[:80], body_style)],
        ["Período analizado:", Paragraph(anio_str, body_style)],
        ["Total de artículos:", str(len(df_art))],
        ["Números analizados:", str(df_art["issue"].nunique())],
        ["Fecha del informe:", time.strftime("%d de %B de %Y, %H:%M")],
    ]
    t_meta = Table(meta_data, colWidths=[4.5*rcm, 13.1*rcm])
    t_meta.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("TEXTCOLOR", (0,0), (0,-1), c_azul_oscuro),
        ("TEXTCOLOR", (1,0), (1,-1), c_negro),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [c_blanco, c_azul_palido]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BCD9F0")),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(t_meta)
    story.append(PageBreak())

    # ── 1. ARTÍCULOS POR NÚMERO ───────────────────────────
    story.append(section_header("1. Artículos por Número (Issue)"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "La siguiente gráfica muestra la cantidad de artículos publicados en cada número de la revista. "
        "Permite identificar si la producción es homogénea entre números o si existen ediciones con "
        "mayor o menor carga de contenido. Números con muy pocos artículos pueden indicar ediciones "
        "especiales, suplementos o períodos de baja producción.",
        body_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(RLImage(io.BytesIO(resultados["fig_issues_bytes"]), width=17*rcm, height=5.5*rcm))
    story.append(Paragraph("Fig. 1 — Distribución de artículos por número de la revista.", caption_style))
    story.append(PageBreak())

    # ── 2. ANÁLISIS TEMPORAL ─────────────────────────────
    story.append(section_header("2. Evolución Temporal de la Producción"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "El análisis temporal permite observar cómo ha variado la producción científica de la revista "
        "a lo largo del tiempo. Se identifican tendencias de crecimiento, estancamiento o declive, "
        "así como el ritmo de adopción de estándares de metadatos (DOI, ORCID, resúmenes, palabras clave). "
        "Un aumento en la cobertura de DOI y ORCID refleja la modernización editorial de la revista.",
        body_style))
    story.append(Spacer(1, 0.1*inch))

    figs_temp = resultados.get("figs_temporales", {})
    if "conteo_anio" in figs_temp:
        conteo_a = figs_temp["conteo_anio"]
        col_anio_k = [c for c in conteo_a.columns if "anio" in c][0]

        fig_temp_mpl, axes = plt.subplots(1, 2, figsize=(14, 4))
        # Barras por año
        axes[0].bar(conteo_a[col_anio_k].astype(str), conteo_a["articulos"],
                    color=PALETTE["600"], edgecolor="white")
        axes[0].set_title("Artículos por año", fontweight="bold", color="#0a0a0a")
        axes[0].set_xlabel("Año")
        axes[0].set_ylabel("Artículos")
        axes[0].tick_params(axis="x", rotation=45)
        axes[0].set_facecolor(PALETTE["50"])
        axes[0].spines["top"].set_visible(False)
        axes[0].spines["right"].set_visible(False)

        # Líneas de evolución
        for col_k, label, color in [
            ("con_doi", "Con DOI", PALETTE["600"]),
            ("con_orcid", "Con ORCID", PALETTE["teal"]),
            ("con_abstract", "Abstract", PALETTE["300"]),
        ]:
            if col_k in conteo_a.columns:
                axes[1].plot(conteo_a[col_anio_k].astype(str), conteo_a[col_k],
                             label=label, color=color, marker="o", linewidth=2)
        axes[1].set_title("Evolución de metadatos", fontweight="bold", color="#0a0a0a")
        axes[1].set_xlabel("Año")
        axes[1].set_ylabel("Artículos")
        axes[1].legend(fontsize=8)
        axes[1].tick_params(axis="x", rotation=45)
        axes[1].set_facecolor(PALETTE["50"])
        axes[1].spines["top"].set_visible(False)
        axes[1].spines["right"].set_visible(False)

        plt.tight_layout()
        story.append(RLImage(io.BytesIO(fig_to_bytes(fig_temp_mpl)), width=17*rcm, height=7*rcm))
        plt.close(fig_temp_mpl)
        story.append(Paragraph(
            "Fig. 2 — Izquierda: producción por año. Derecha: adopción de metadatos estructurados.",
            caption_style))
    story.append(PageBreak())

    # ── 3. METADATOS ──────────────────────────────────────
    story.append(section_header("3. Presencia de Metadatos Estructurados"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "Los metadatos estructurados son esenciales para la visibilidad, indexación y citación de "
        "los artículos en bases de datos internacionales. El DOI (Digital Object Identifier) garantiza "
        "un enlace permanente; el ORCID vincula el trabajo a un perfil de investigador único; "
        "las palabras clave y resúmenes facilitan la recuperación en motores de búsqueda académicos. "
        "Una cobertura superior al 70% se considera buena práctica editorial.",
        body_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(RLImage(io.BytesIO(resultados["fig_metadatos"]), width=17*rcm, height=8*rcm))
    story.append(Paragraph(
        "Fig. 3 — Porcentaje de artículos con cada tipo de metadato. La línea vertical marca el 50%.",
        caption_style))
    story.append(PageBreak())

    # ── 4. AUTORES ────────────────────────────────────────
    story.append(section_header("4. Análisis de Autores"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "El análisis de autoría identifica a los investigadores más productivos en la revista. "
        "Autores con múltiples publicaciones son considerados autores nucleares o recurrentes, "
        "lo que puede indicar una línea editorial temática consistente. La detección de ORCID "
        "permite vincular a estos autores con sus perfiles institucionales y otras publicaciones.",
        body_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(RLImage(io.BytesIO(resultados["fig_autores"]), width=17*rcm, height=7*rcm))
    story.append(Paragraph(
        "Fig. 4 — Top 20 autores con mayor número de artículos publicados en el período analizado.",
        caption_style))
    story.append(PageBreak())

    # ── 5. MAPA GEOGRÁFICO ────────────────────────────────
    story.append(section_header("5. Distribución Geográfica de Autores"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "La distribución geográfica de autores refleja el alcance y la internacionalización de la "
        "revista. Se detecta a partir de las afiliaciones institucionales declaradas en los metadatos "
        "de cada artículo. Revistas con mayor diversidad geográfica suelen tener mayor impacto y "
        "visibilidad internacional. Las líneas en el mapa interactivo representan colaboraciones "
        "internacionales (co-autorías entre autores de distintos países en el mismo artículo).",
        body_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "⚠ Nota: El mapa interactivo con burbujas y conexiones está disponible en la aplicación web. "
        "Este informe muestra la tabla resumen de países detectados.",
        Paragraph("", body_style).__class__(
            "⚠ Nota: El mapa interactivo está disponible en la aplicación web.",
            ParagraphStyle("note", fontSize=8.5, textColor=c_azul_medio, fontName="Helvetica-Oblique")
        ) if False else body_style))

    df_mapa = resultados.get("df_mapa", pd.DataFrame())
    if not df_mapa.empty:
        paises_count = Counter(df_mapa["pais"].tolist())
        total_geo = sum(paises_count.values())

        # Gráfica matplotlib de países
        top_paises = paises_count.most_common(15)
        paises_names = [p[0] for p in top_paises]
        paises_vals = [p[1] for p in top_paises]
        fig_p, ax_p = plt.subplots(figsize=(11, 5))
        colors_p = [PALETTE["600"] if i == 0 else PALETTE["300"] if i < 3 else PALETTE["100"]
                    for i in range(len(paises_names))]
        ax_p.barh(paises_names, paises_vals, color=colors_p, edgecolor="white")
        ax_p.set_title("Top países por número de autores", fontweight="bold", color="#0a0a0a")
        ax_p.set_xlabel("Autores")
        ax_p.invert_yaxis()
        ax_p.set_facecolor(PALETTE["50"])
        ax_p.spines["top"].set_visible(False)
        ax_p.spines["right"].set_visible(False)
        for i, v in enumerate(paises_vals):
            ax_p.text(v + 0.1, i, f" {v} ({round(v/total_geo*100,1)}%)", va="center", fontsize=8)
        plt.tight_layout()
        story.append(RLImage(io.BytesIO(fig_to_bytes(fig_p)), width=17*rcm, height=7*rcm))
        plt.close(fig_p)
        story.append(Paragraph("Fig. 5 — Países con mayor presencia de autores.", caption_style))
    story.append(PageBreak())

    # ── 6. CLUSTERS TEMÁTICOS ─────────────────────────────
    story.append(section_header("6. Clustering Temático (KMeans + TF-IDF)"))
    story.append(Spacer(1, 0.1*inch))

    # Explicación detallada
    explain_text = """
<b>¿Cómo se obtienen los clusters?</b><br/><br/>

<b>Paso 1 — Matriz TF-IDF:</b> El texto de cada artículo (título + resumen + palabras clave) se convierte
en una representación numérica usando TF-IDF (Term Frequency–Inverse Document Frequency):<br/>
<i>TF-IDF(término, documento) = TF × IDF</i><br/>
Donde <b>TF</b> mide cuánto aparece un término en el artículo y <b>IDF</b> mide cuán raro es en toda la
colección. Las palabras muy comunes (como "investigación") pierden peso; las específicas
(como "sismología") ganan importancia. Se filtran palabras académicas genéricas (artículo, libro,
estudio, etc.) para que los términos sean más informativos.<br/><br/>

<b>Paso 2 — K-Means:</b> Es un algoritmo de aprendizaje no supervisado que agrupa artículos por
cercanía en el espacio TF-IDF. Cada artículo se asigna al centroide (punto central) más cercano
usando distancia euclídea. Los grupos resultantes son <b>clusters temáticos</b>, donde los artículos
dentro de cada cluster son más similares entre sí que con los de otros clusters.<br/><br/>

<b>Paso 3 — PCA:</b> Como la matriz TF-IDF tiene miles de dimensiones, se usa el Análisis de
Componentes Principales (PCA) para reducirla a 2 dimensiones (X,Y) y poder visualizar los clusters
en un plano 2D.
"""
    story.append(Paragraph(explain_text,
        ParagraphStyle("explain", parent=styles["Normal"],
            fontSize=8.5, leading=13, textColor=c_negro,
            backColor=c_azul_palido, borderPadding=10,
            leftIndent=6, rightIndent=6, spaceBefore=4, spaceAfter=8)))
    story.append(Spacer(1, 0.1*inch))

    img_row = Table(
        [[RLImage(io.BytesIO(resultados["fig_clusters"]), width=8*rcm, height=5.5*rcm),
          RLImage(io.BytesIO(resultados["fig_pca"]), width=9*rcm, height=7*rcm)]],
        colWidths=[8.5*rcm, 9.1*rcm]
    )
    story.append(img_row)
    story.append(Paragraph(
        "Fig. 6a — Artículos por cluster temático. Fig. 6b — Proyección PCA: cada punto es un artículo.",
        caption_style))
    story.append(PageBreak())

    # ── 7. TÓPICOS LDA ────────────────────────────────────
    story.append(section_header("7. Modelado de Tópicos (LDA)"))
    story.append(Spacer(1, 0.1*inch))

    explain_lda = """
<b>¿Cómo funciona LDA?</b><br/><br/>

A diferencia de los clusters (donde cada artículo pertenece a un solo grupo), LDA
(Latent Dirichlet Allocation) es un <b>modelo probabilístico</b> que asume:<br/>
• Cada artículo es una mezcla de tópicos (ej: 70% "Geofísica" + 30% "Historia").<br/>
• Cada tópico es una distribución de palabras con distintas probabilidades.<br/><br/>

El modelo detecta automáticamente estas estructuras latentes al observar qué palabras
aparecen juntas en los mismos documentos. A cada artículo se le asigna el tópico con
mayor probabilidad (<i>topic_values.argmax()</i>).<br/><br/>

<b>Tópico regional:</b> Se detecta automáticamente cuando un tópico contiene 2 o más términos
geográficos o de escala territorial (país, región, municipal, zona, comunidad, etc.).
"""
    story.append(Paragraph(explain_lda,
        ParagraphStyle("explain_lda", parent=styles["Normal"],
            fontSize=8.5, leading=13, textColor=c_negro,
            backColor=c_azul_palido, borderPadding=10,
            leftIndent=6, rightIndent=6, spaceBefore=4, spaceAfter=8)))

    story.append(RLImage(io.BytesIO(resultados["fig_topicos_issue"]), width=17*rcm, height=6.5*rcm))
    story.append(Paragraph(
        "Fig. 7 — Distribución de tópicos por número de la revista. 🌍 = tópico con orientación regional.",
        caption_style))
    story.append(Spacer(1, 0.1*inch))

    wc_figs = resultados["wc_figs"]
    topicos_reg = resultados.get("topicos_regionales", {})
    for i in range(0, len(wc_figs), 2):
        row_imgs = []
        for j in [i, i+1]:
            if j < len(wc_figs) and wc_figs[j]:
                row_imgs.append(RLImage(io.BytesIO(wc_figs[j]), width=8.2*rcm, height=4.1*rcm))
            else:
                row_imgs.append(Spacer(1, 1))
        story.append(Table([row_imgs], colWidths=[8.8*rcm, 8.8*rcm]))
    story.append(Paragraph(
        "Fig. 7b — Nubes de palabras por tópico. Las palabras más grandes tienen mayor peso en el tópico.",
        caption_style))
    story.append(PageBreak())

    # ── 8. RED DE COAUTORÍA ───────────────────────────────
    story.append(section_header("8. Red de Coautoría"))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "La red de coautoría representa las relaciones de colaboración entre autores. Cada nodo es "
        "un autor y cada arista conecta a dos autores que han publicado juntos. El tamaño del nodo "
        "refleja el número de artículos del autor. El color indica el tópico dominante de su producción. "
        "Los nodos con el ícono 🔗 tienen ORCID registrado, lo que permite verificar su identidad "
        "y producción en plataformas internacionales. En la aplicación web puedes hacer doble clic "
        "sobre cualquier nodo con ORCID para abrir su perfil directamente.",
        body_style))
    story.append(Spacer(1, 0.1*inch))

    G = resultados.get("grafo_G", nx.Graph())
    con_orcid = sum(1 for n in G.nodes() if G.nodes[n].get("orcid") or orcid_map.get(n, ""))
    metricas_red = [
        ["Nodos (autores en red)", str(G.number_of_nodes())],
        ["Aristas (colaboraciones)", str(G.number_of_edges())],
        ["Densidad de la red", f"{nx.density(G):.4f}" if G.number_of_nodes() > 1 else "—"],
        ["Componentes conectados", str(nx.number_connected_components(G)) if G.number_of_nodes() > 0 else "—"],
        ["Autores con ORCID", str(con_orcid)],
        ["Países en la red", str(len(set(G.nodes[n].get("pais","") for n in G.nodes() if G.nodes[n].get("pais"))))],
    ]

    t_red_pdf = Table(
        [["Métrica", "Valor"]] + metricas_red,
        colWidths=[9*rcm, 8.6*rcm]
    )
    t_red_pdf.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), c_azul_oscuro),
        ("TEXTCOLOR", (0,0), (-1,0), c_blanco),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("TEXTCOLOR", (0,1), (-1,-1), c_negro),
        ("TEXTCOLOR", (0,1), (0,-1), c_azul_oscuro),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [c_blanco, c_azul_palido]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BCD9F0")),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("ALIGN", (1,0), (1,-1), "CENTER"),
    ]))
    story.append(t_red_pdf)
    story.append(Spacer(1, 0.12*inch))

    if resultados["fig_red"]:
        story.append(RLImage(io.BytesIO(resultados["fig_red"]), width=17*rcm, height=13*rcm))
        story.append(Paragraph(
            "Fig. 8 — Red de coautoría. Tamaño = número de artículos. Color = tópico dominante. "
            "Solo se etiquetan autores con 2 o más artículos.",
            caption_style))
    else:
        story.append(Paragraph(
            "No se encontraron pares de coautoría en el período seleccionado.", body_style))

    story.append(Spacer(1, 0.3*inch))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#BCD9F0")))
    story.append(Spacer(1, 0.08*inch))
    story.append(Paragraph(
        f"Analizador OJS v5.0  ·  Paleta azul  ·  "
        f"Generado el {time.strftime('%d/%m/%Y a las %H:%M')}",
        center_style,
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{PALETTE['800']},{PALETTE['700']});
                padding:1rem;border-radius:10px;border:1px solid {PALETTE['500']}40;
                margin-bottom:1rem;">
      <h3 style="color:{PALETTE['400']};margin:0;font-size:1rem;">⚙️ Configuración</h3>
    </div>
    """, unsafe_allow_html=True)

    url_input = st.text_input(
        "🔗 URL del archivo de números",
        placeholder="https://ejemplo.com/index.php/revista/issue/archive",
    )
    st.markdown("##### Parámetros de análisis")
    n_clusters = st.slider("Clusters KMeans", 2, 10, 4)
    n_topics   = st.slider("Tópicos LDA", 2, 10, 5)

    st.markdown("---")
    st.markdown(f"""
    <div style="background:{PALETTE['800']};border-radius:8px;padding:0.8rem;
                border:1px solid {PALETTE['700']};">
      <p style="color:{PALETTE['300']};font-size:0.8rem;margin:0;font-weight:600;">Flujo de trabajo</p>
      <ol style="color:{PALETTE['200']};font-size:0.78rem;padding-left:1rem;margin-top:0.5rem;">
        <li>Pega la URL del archivo de la revista</li>
        <li><b>Cargar revista</b> → visualiza números</li>
        <li>Filtra y selecciona cuántos analizar</li>
        <li><b>Extraer artículos</b> (metadatos completos)</li>
        <li>Usa el <b>selector de período</b> dinámico</li>
        <li><b>Ejecutar análisis</b> en el período elegido</li>
        <li>Explora el mapa, la red y descarga PDF + Excel</li>
      </ol>
    </div>
    """, unsafe_allow_html=True)
    st.caption(f"OJS 2.x / 3.x · v5.0 · Paleta azul")


# ═══════════════════════════════════════════════════════
# STATE INIT
# ═══════════════════════════════════════════════════════
for key in ["df_issues", "df_articulos", "resultados", "url_usada"]:
    if key not in st.session_state:
        st.session_state[key] = None if key != "url_usada" else ""

stop_words = load_stopwords()


# ═══════════════════════════════════════════════════════
# PASO 1: CARGAR ISSUES
# ═══════════════════════════════════════════════════════
cargar_btn = st.button(
    "🔍 Cargar revista (todos los números)", type="primary",
    use_container_width=True, disabled=not url_input.strip()
)

if cargar_btn and url_input.strip():
    with st.spinner("Recorriendo todas las páginas de la revista…"):
        df_issues, err = extraer_issues_todas_paginas(url_input.strip())
    if err:
        st.error(f"❌ Error: {err}")
    elif df_issues.empty:
        st.warning("⚠️ No se encontraron números. Verifica la URL.")
    else:
        st.session_state.df_issues = df_issues
        st.session_state.url_usada = url_input.strip()
        st.session_state.df_articulos = None
        st.session_state.resultados = None
        st.success(f"Se encontraron **{len(df_issues)}** números.")


# ═══════════════════════════════════════════════════════
# PASO 2: VISUALIZAR Y FILTRAR ISSUES
# ═══════════════════════════════════════════════════════
if st.session_state.df_issues is not None:
    df_issues = st.session_state.df_issues
    st.markdown('<div class="section-header"> Números disponibles en la revista</div>',
                unsafe_allow_html=True)

    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    metricas_issues = [
        (len(df_issues), "Total de números", ""),
        (df_issues["anio"].replace("", pd.NA).dropna().nunique(), "Años distintos", "accent"),
        (df_issues["volumen"].replace("", pd.NA).dropna().nunique(), "Volúmenes", "teal"),
        (df_issues["tipo"].value_counts().index[0] if len(df_issues) > 0 else "—", "Tipo dominante", "dark"),
    ]
    for col, (val, label, cls) in zip([col_m1, col_m2, col_m3, col_m4], metricas_issues):
        with col:
            st.markdown(f'<div class="metric-card {cls}"><h3>{val}</h3><p>{label}</p></div>',
                        unsafe_allow_html=True)

    with st.expander("Filtrar números por tipo / año / volumen", expanded=False):
        fcol1, fcol2, fcol3 = st.columns(3)
        with fcol1:
            tipos_unicos = ["Todos"] + sorted(df_issues["tipo"].unique().tolist())
            filtro_tipo = st.selectbox("Tipo", tipos_unicos, key="ftipo")
        with fcol2:
            anios_unicos = ["Todos"] + sorted([a for a in df_issues["anio"].unique() if a], reverse=True)
            filtro_anio = st.selectbox("Año", anios_unicos, key="fanio")
        with fcol3:
            vols_unicos = ["Todos"] + sorted([v for v in df_issues["volumen"].unique() if v])
            filtro_vol = st.selectbox("Volumen", vols_unicos, key="fvol")

        df_filtrado = df_issues.copy()
        if filtro_tipo != "Todos":
            df_filtrado = df_filtrado[df_filtrado["tipo"] == filtro_tipo]
        if filtro_anio != "Todos":
            df_filtrado = df_filtrado[df_filtrado["anio"] == filtro_anio]
        if filtro_vol != "Todos":
            df_filtrado = df_filtrado[df_filtrado["volumen"] == filtro_vol]
        st.info(f"Mostrando {len(df_filtrado)} de {len(df_issues)} números")

    if "df_filtrado" not in dir():
        df_filtrado = df_issues.copy()

    cols_mostrar = [c for c in ["issue", "tipo", "anio", "volumen", "numero", "issue_url"]
                    if c in df_filtrado.columns]
    st.dataframe(
        df_filtrado[cols_mostrar].rename(columns={
            "issue": "Título", "tipo": "Tipo", "anio": "Año",
            "volumen": "Vol.", "numero": "Núm.", "issue_url": "URL"
        }),
        use_container_width=True, height=260, hide_index=True
    )

    st.markdown('<div class="section-header">📥 Extracción de artículos</div>', unsafe_allow_html=True)
    col_n1, col_n2 = st.columns([2, 1])
    with col_n1:
        usar_filtrado = st.checkbox("Usar solo los números filtrados", value=False)
        df_para_extraer = df_filtrado if usar_filtrado else df_issues
    with col_n2:
        n_analizar = st.number_input(
            "¿Cuántos números analizar?",
            min_value=1, max_value=len(df_para_extraer),
            value=min(3, len(df_para_extraer)),
        )

    extraer_btn = st.button(
        "📥 Extraer artículos (metadatos completos + ORCID)",
        type="primary", use_container_width=True
    )

    if extraer_btn:
        df_sel = df_para_extraer.head(int(n_analizar))
        todos = []
        prog = st.progress(0, text="Iniciando extracción…")
        status = st.empty()
        for i, (_, row) in enumerate(df_sel.iterrows()):
            issue_meta = row.to_dict()
            status.info(f"📖 Procesando: **{row['issue']}**")
            arts = extraer_articulos_issue_completo(row["issue_url"], issue_meta)
            todos.extend(arts)
            prog.progress((i + 1) / len(df_sel),
                          text=f"Número {i+1}/{len(df_sel)} — {len(arts)} artículos")
            time.sleep(0.3)
        prog.empty(); status.empty()
        if todos:
            df_art = pd.DataFrame(todos).drop_duplicates(subset=["articulo_url"])
            st.session_state.df_articulos = df_art
            st.session_state.resultados = None
            st.success(f"{len(df_art)} artículos extraídos de {int(n_analizar)} números.")
        else:
            st.error("No se pudieron extraer artículos.")


# ═══════════════════════════════════════════════════════
# PASO 3: VISUALIZAR ARTÍCULOS + SELECTOR DE PERÍODO
# ═══════════════════════════════════════════════════════
if st.session_state.df_articulos is not None:
    df_art_completo = st.session_state.df_articulos

    st.markdown('<div class="section-header">📄 Artículos extraídos · Selección de período</div>',
                unsafe_allow_html=True)

    df_periodo, anio_ini, anio_fin = widget_selector_periodo(df_art_completo, key_prefix="main_period")

    col_m = st.columns(4)
    c_doi   = df_periodo["tiene_doi"].fillna(False).astype(bool).sum() if "tiene_doi" in df_periodo.columns else 0
    c_orcid = df_periodo["tiene_orcid"].fillna(False).astype(bool).sum() if "tiene_orcid" in df_periodo.columns else 0

    metricas_periodo = [
        (len(df_periodo), "Artículos en período", ""),
        (df_periodo["autores"].dropna().apply(lambda x: [a.strip() for a in x.split(",") if a.strip()])
         .explode().nunique() if not df_periodo.empty else 0, "Autores únicos", "accent"),
        (c_doi, "Con DOI", "teal"),
        (c_orcid, "Con ORCID", "dark"),
    ]
    for col, (val, label, cls) in zip(col_m, metricas_periodo):
        with col:
            st.markdown(f'<div class="metric-card {cls}"><h3>{val}</h3><p>{label}</p></div>',
                        unsafe_allow_html=True)

    col_m2 = st.columns(4)
    c_abs  = df_periodo["tiene_abstract"].fillna(False).astype(bool).sum() if "tiene_abstract" in df_periodo.columns else 0
    c_kw   = df_periodo["tiene_keywords"].fillna(False).astype(bool).sum() if "tiene_keywords" in df_periodo.columns else 0
    c_afil = df_periodo["tiene_afiliacion"].fillna(False).astype(bool).sum() if "tiene_afiliacion" in df_periodo.columns else 0
    c_pais = df_periodo["tiene_pais"].fillna(False).astype(bool).sum() if "tiene_pais" in df_periodo.columns else 0

    for col, (val, label, cls) in zip(col_m2, [
        (c_abs, "Con Abstract", "teal"),
        (c_kw, "Con Keywords", ""),
        (c_afil, "Con Afiliación", "accent"),
        (c_pais, "País detectado", "dark"),
    ]):
        with col:
            st.markdown(f'<div class="metric-card {cls}"><h3>{val}</h3><p>{label}</p></div>',
                        unsafe_allow_html=True)

    with st.expander("Ver tabla de artículos del período seleccionado", expanded=False):
        cols_tabla = [c for c in [
            "issue", "anio_issue", "volumen_issue", "numero_issue",
            "articulo", "autores", "doi", "keywords",
            "tiene_doi", "tiene_orcid", "tiene_abstract", "tiene_keywords",
        ] if c in df_periodo.columns]
        st.dataframe(df_periodo[cols_tabla], use_container_width=True, height=320, hide_index=True)

    st.markdown('<div class="section-header">🔬 Ejecutar análisis de contenido</div>',
                unsafe_allow_html=True)

    if len(df_periodo) < 2:
        st.warning("⚠️ Selecciona al menos 2 artículos para ejecutar el análisis.")
    else:
        periodo_label = f"{anio_ini}–{anio_fin}" if anio_ini and anio_fin else "período completo"
        analizar_btn = st.button(
            f"🚀 Ejecutar análisis · {len(df_periodo)} artículos · {periodo_label}",
            type="primary", use_container_width=True
        )

        if analizar_btn:
            prog2 = st.progress(0, text="Iniciando…")
            status2 = st.empty()

            def cb(pct, msg):
                prog2.progress(pct, text=msg)
                status2.info(msg)

            res = analizar_datos(df_periodo.copy(), n_clusters, n_topics, stop_words, cb)
            res["periodo_label"] = periodo_label
            res["anio_ini"] = anio_ini
            res["anio_fin"] = anio_fin
            st.session_state.resultados = res
            if "df_articulos" in res:
                st.session_state.df_analizado = res["df_articulos"]
            prog2.empty(); status2.empty()
            st.success(f"✅ Análisis completado para el período {periodo_label}.")


# ═══════════════════════════════════════════════════════
# PASO 4: RESULTADOS
# ═══════════════════════════════════════════════════════
if st.session_state.resultados is not None:
    res = st.session_state.resultados
    df_final = res.get("df_articulos", st.session_state.df_articulos)
    url_usada = st.session_state.url_usada
    periodo_label = res.get("periodo_label", "")
    anio_ini = res.get("anio_ini")
    anio_fin = res.get("anio_fin")
    orcid_map = res.get("orcid_map", {})

    st.markdown("---")
    st.markdown(f"""
    <div class="section-header">
        📊 Resultados del análisis
        {'· <span style="font-size:0.85rem;opacity:0.7;">Período: ' + periodo_label + '</span>' if periodo_label else ''}
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "Temporal", "Números", "Metadatos",
        "Autores", "Mapa", "Clusters", "Tópicos", "Red"
    ])

    # ── Tab 1: TEMPORAL ───────────────────────────────────
    with tab1:
        st.subheader(f"Análisis temporal · {periodo_label}")
        if st.session_state.df_articulos is not None:
            df_refilter, rf_ini, rf_fin = widget_selector_periodo(
                st.session_state.df_articulos, key_prefix="tab1_period"
            )
            if len(df_refilter) > 0:
                figs_rf = generar_analisis_temporal(df_refilter)
                for fig_key in ["fig_anio", "fig_meta_anio", "fig_pct_anio", "fig_heatmap"]:
                    if fig_key in figs_rf:
                        st.plotly_chart(figs_rf[fig_key], use_container_width=True)
                if "conteo_anio" in figs_rf:
                    st.subheader("Tabla resumen por año")
                    st.dataframe(figs_rf["conteo_anio"], use_container_width=True, hide_index=True)
                if "fig_vol" in figs_rf:
                    st.plotly_chart(figs_rf["fig_vol"], use_container_width=True)
        else:
            figs_temp = res.get("figs_temporales", {})
            for fig_key in ["fig_anio", "fig_meta_anio", "fig_pct_anio", "fig_heatmap"]:
                if fig_key in figs_temp:
                    st.plotly_chart(figs_temp[fig_key], use_container_width=True)

    # ── Tab 2: NÚMEROS ────────────────────────────────────
    with tab2:
        st.subheader("Artículos por número (issue)")
        if st.session_state.df_articulos is not None:
            df_refilter2, _, _ = widget_selector_periodo(
                st.session_state.df_articulos, key_prefix="tab2_period"
            )
            if not df_refilter2.empty:
                conteo_iss = df_refilter2.groupby("issue").size().reset_index(name="num_articulos")
                conteo_iss = conteo_iss.sort_values("num_articulos", ascending=False)
                fig_iss = go.Figure(go.Bar(
                    x=conteo_iss["issue"].apply(lambda t: t[:35] + "…" if len(t) > 35 else t),
                    y=conteo_iss["num_articulos"],
                    marker=dict(color=conteo_iss["num_articulos"], colorscale=BLUE_SCALE),
                    text=conteo_iss["num_articulos"], textposition="outside",
                ))
                fig_iss.update_layout(
                    xaxis=dict(title="", tickangle=-50, type="category"),
                    yaxis=dict(title="Artículos"),
                    plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
                    height=420, font=dict(color="#0a0a0a"),
                )
                st.plotly_chart(fig_iss, use_container_width=True)
                st.dataframe(conteo_iss, use_container_width=True, hide_index=True)
        else:
            st.plotly_chart(res.get("fig_issues_plotly", go.Figure()), use_container_width=True)

    # ── Tab 3: METADATOS ──────────────────────────────────
    with tab3:
        st.subheader("Presencia de metadatos")
        if st.session_state.df_articulos is not None:
            df_refilter3, _, _ = widget_selector_periodo(
                st.session_state.df_articulos, key_prefix="tab3_period"
            )
            if not df_refilter3.empty:
                fig_meta_din = grafica_metadatos_presencia(df_refilter3, f"n={len(df_refilter3)}")
                st.plotly_chart(fig_meta_din, use_container_width=True)
        else:
            st.plotly_chart(res.get("fig_metadatos_plotly", go.Figure()), use_container_width=True)

        meta_cols = [c for c in [
            "articulo", "autores", "doi", "keywords", "tiene_doi",
            "tiene_orcid", "tiene_abstract", "tiene_keywords", "tiene_afiliacion", "tiene_pais",
        ] if c in df_final.columns]
        st.subheader("Detalle por artículo")
        st.dataframe(df_final[meta_cols], use_container_width=True, height=380, hide_index=True)

    # ── Tab 4: AUTORES ────────────────────────────────────
    with tab4:
        st.subheader("Análisis de autores")

        st.markdown(f"""
        <div style="background-color: {PALETTE['50']}; padding: 10px; border-radius: 5px; margin-bottom: 20px; border-left: 4px solid {PALETTE['600']};">
        <b>Control de visualización:</b> Ajusta los autores en la nube. Límite máximo establecido en 100 para proteger el rendimiento de la aplicación.
        </div>
        """, unsafe_allow_html=True)
        max_autores_nube = st.slider("Cantidad de autores en la nube", 10, 100, 40, key="slider_nube")

        if st.session_state.df_articulos is not None:
            df_refilter4, _, _ = widget_selector_periodo(
                st.session_state.df_articulos, key_prefix="tab4_period"
            )
            if not df_refilter4.empty:
                todos_aut = []
                for x in df_refilter4["autores"].dropna():
                    todos_aut.extend([a.strip() for a in x.split(",") if a.strip()])
                cont_aut = Counter(todos_aut)
                df_aut_din = (pd.DataFrame(cont_aut.items(), columns=["autor", "num_articulos"])
                              .sort_values("num_articulos", ascending=False)
                              .reset_index(drop=True))
                df_aut_din["ORCID"] = df_aut_din["autor"].map(lambda a: orcid_map.get(a, ""))

                col_nube, col_bar = st.columns([1, 1])

                with col_nube:
                    st.markdown("#### Nube de Palabras")
                    img_nube = generar_nube_autores_bytes(df_aut_din, max_autores_nube)
                    if img_nube:
                        st.image(img_nube, caption=f"Nube de palabras (Top {max_autores_nube})", use_container_width=True)
                    else:
                        st.info("Sin datos suficientes para generar la nube.")

                with col_bar:
                    st.metric("Autores únicos en período", len(df_aut_din))
                    top20_din = df_aut_din.head(20)
                    fig_aut_din = go.Figure(go.Bar(
                        x=top20_din["num_articulos"], y=top20_din["autor"],
                        orientation="h",
                        marker=dict(color=top20_din["num_articulos"], colorscale=BLUE_SCALE, line=dict(width=0)),
                        text=top20_din["num_articulos"], textposition="outside",
                    ))
                    fig_aut_din.update_layout(
                        title=dict(text="Top 20 autores por producción", font=dict(size=15, color="#0a0a0a")),
                        xaxis=dict(title="Artículos"),
                        yaxis=dict(title="", autorange="reversed"),
                        plot_bgcolor=PALETTE["50"], paper_bgcolor="white",
                        height=400, margin=dict(l=150, r=40, t=40, b=40),
                        font=dict(color="#0a0a0a"),
                    )
                    st.plotly_chart(fig_aut_din, use_container_width=True)

                st.subheader("Directorio detallado")
                st.dataframe(
                    df_aut_din.head(50), use_container_width=True, hide_index=True, height=380,
                    column_config={"ORCID": st.column_config.LinkColumn("ORCID 🔗", display_text="Ver perfil")}
                )
        else:
            df_aut = res["df_autores_count"].copy()
            df_aut["ORCID"] = df_aut["autor"].map(lambda a: orcid_map.get(a, ""))

            col_nube, col_bar = st.columns([1, 1])

            with col_nube:
                st.markdown("#### Nube de Palabras")
                img_nube_stat = generar_nube_autores_bytes(df_aut, max_autores_nube)
                if img_nube_stat:
                    st.image(img_nube_stat, caption=f"Nube de palabras (Top {max_autores_nube})", use_container_width=True)

            with col_bar:
                st.metric("Autores únicos en total", len(df_aut))
                st.plotly_chart(res.get("fig_autores_plotly", go.Figure()), use_container_width=True)

            st.subheader("Directorio detallado")
            st.dataframe(df_aut.head(50), use_container_width=True, hide_index=True,
                         column_config={"ORCID": st.column_config.LinkColumn("ORCID 🔗", display_text="Ver perfil")})

    # ── Tab 5: MAPA ───────────────────────────────────────
    with tab5:
        st.subheader("🌍 Mapa geográfico de autores")
        if st.session_state.df_articulos is not None:
            df_refilter5, _, _ = widget_selector_periodo(
                st.session_state.df_articulos, key_prefix="tab5_period"
            )
            if not df_refilter5.empty:
                fig_mapa_din, df_mapa_din = generar_mapa_autores(df_refilter5)
                if fig_mapa_din is not None:
                    st.markdown("""
                    <div class="network-controls">
                    💡 <b>Burbujas</b> = países (tamaño proporcional a autores).
                    <b>Líneas</b> = colaboraciones internacionales en el mismo artículo.
                    </div>
                    """, unsafe_allow_html=True)
                    st.plotly_chart(fig_mapa_din, use_container_width=True)
                    fig_paises_din = generar_grafica_paises(df_refilter5)
                    if fig_paises_din:
                        st.plotly_chart(fig_paises_din, use_container_width=True)
                    if not df_mapa_din.empty:
                        st.subheader("Resumen por país")
                        df_pais_res = (df_mapa_din.groupby("pais")
                                       .agg(autores=("autor", "count"),
                                            articulos=("articulo", "nunique"))
                                       .reset_index()
                                       .sort_values("autores", ascending=False))
                        st.dataframe(df_pais_res, use_container_width=True, hide_index=True)
                else:
                    st.info("No se detectaron países en este período.")
        else:
            fig_mapa = res.get("fig_mapa")
            if fig_mapa:
                st.plotly_chart(fig_mapa, use_container_width=True)

    # ── Tab 6: CLUSTERS ───────────────────────────────────
    with tab6:
        st.subheader(f"Clustering Temático · {periodo_label}")

        st.markdown("""
        <div class="explanation-box">
        <h4>🔬 ¿Cómo se calculan los clusters?</h4>
        <p><b>1. TF-IDF:</b> Cada artículo (título + resumen + keywords) se convierte en un vector numérico.
        Las palabras raras y específicas (ej: "sismología") tienen mayor peso que las comunes (ej: "estudio").</p>
        <p><b>2. K-Means:</b> Agrupa los artículos en k clusters según la distancia euclídea entre sus vectores.
        Los artículos dentro de un mismo cluster son más similares entre sí que con los de otros clusters.</p>
        <p><b>3. PCA:</b> Reduce las miles de dimensiones TF-IDF a solo 2 (X,Y) para poder visualizarlos en una gráfica.</p>
        <p><b>Filtro activo:</b> Se excluyen palabras genéricas como "artículo", "libro", "estudio", "investigación", etc.</p>
        </div>
        """, unsafe_allow_html=True)

        col_c1, col_c2 = st.columns([1, 2])
        with col_c1:
            st.dataframe(res["conteo_clusters"], use_container_width=True, hide_index=True)
            for c, terms in res["top_terms_clusters"].items():
                st.markdown(f"**Cluster {c}:** {', '.join(terms[:8])}")
        with col_c2:
            st.plotly_chart(res.get("fig_pca_plotly", go.Figure()), use_container_width=True)
        st.plotly_chart(res.get("fig_clusters_plotly", go.Figure()), use_container_width=True)

    # ── Tab 7: TÓPICOS ────────────────────────────────────
    with tab7:
        st.subheader(f"Modelado de Tópicos LDA · {periodo_label}")

        topicos_reg = res.get("topicos_regionales", {})

        st.markdown("""
        <div class="explanation-box">
        <h4>🏷️ ¿Cómo funciona LDA?</h4>
        <p><b>LDA</b> (Latent Dirichlet Allocation) es un modelo probabilístico que asume dos cosas:</p>
        <p>• Cada artículo es una <b>mezcla de tópicos</b> (ej: 70% geofísica + 30% historia).</p>
        <p>• Cada tópico es una <b>distribución de palabras</b> (el tópico "sismos" asigna alta probabilidad a palabras como epicentro, magnitud, placa).</p>
        <p>A diferencia de K-Means, LDA no asigna un artículo a un solo grupo, sino que calcula una probabilidad para cada tópico y asigna el de mayor valor.</p>
        <p> <b>Tópico regional:</b> detectado automáticamente cuando el tópico contiene ≥2 términos geográficos.</p>
        </div>
        """, unsafe_allow_html=True)

        if topicos_reg:
            st.success(f" Se detectaron **{len(topicos_reg)}** tópicos con orientación regional: "
                       + ", ".join([f"Tópico {t}" for t in topicos_reg.keys()]))

        for t, terms in res["top_terms_topics"].items():
            regional_badge = "  [REGIONAL]" if t in topicos_reg else ""
            with st.expander(f"Tópico {t}{regional_badge}: {', '.join(terms[:5])}"):
                st.write(", ".join(terms))
                if t in topicos_reg:
                    st.info(f"Términos regionales detectados: {', '.join(topicos_reg[t])}")

        st.plotly_chart(res.get("fig_topicos_plotly", go.Figure()), use_container_width=True)
        st.subheader("Nubes de palabras por tópico")
        cols_wc = st.columns(2)
        for i, wc_bytes in enumerate(res["wc_figs"]):
            if wc_bytes:
                regional = "  " if i in topicos_reg else ""
                with cols_wc[i % 2]:
                    st.image(wc_bytes, caption=f"Tópico {i}{regional}", use_container_width=True)

    # ── Tab 8: RED ────────────────────────────────────────
    with tab8:
        G = res.get("grafo_G", nx.Graph())
        st.subheader(f"Red interactiva de coautoría · {periodo_label}")

        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1: st.metric("Autores en red", G.number_of_nodes())
        with col_r2: st.metric("Colaboraciones", G.number_of_edges())
        with col_r3:
            n_con_orcid = sum(1 for n in G.nodes()
                              if G.nodes[n].get("orcid") or orcid_map.get(n))
            st.metric("Con ORCID 🔗", n_con_orcid)
        with col_r4:
            n_paises_red = len(set(G.nodes[n].get("pais","") for n in G.nodes()
                                   if G.nodes[n].get("pais")))
            st.metric("Países en red ", n_paises_red)

        if G.number_of_edges() > 0:
            st.markdown("""
            <div class="network-controls">
            💡 <b>Clic</b> = detalles y co-autores.
            <b>Doble clic</b> = abrir perfil ORCID.
            <b>🔗 botón</b> = resaltar solo autores con ORCID.
            <b>Búsqueda</b> por nombre, país o institución.
            </div>
            """, unsafe_allow_html=True)
            html_red = generar_red_interactiva_html(G, orcid_map, df_final)
            st.components.v1.html(html_red, height=720, scrolling=False)

            st.subheader("Directorio de autores")
            df_autores_red = pd.DataFrame([{
                "Autor": n,
                "Artículos": G.nodes[n].get("articulos", 0),
                "Tópico": G.nodes[n].get("topico", "—"),
                "Co-autores": G.degree(n),
                "País": G.nodes[n].get("pais", ""),
                "Afiliación": G.nodes[n].get("afiliacion", "")[:60],
                "ORCID URL": G.nodes[n].get("orcid", "") or orcid_map.get(n, ""),
                "ORCID": "" if (G.nodes[n].get("orcid") or orcid_map.get(n)) else "—",
            } for n in G.nodes()]).sort_values("Artículos", ascending=False)

            st.dataframe(
                df_autores_red, use_container_width=True, height=380,
                column_config={
                    "ORCID URL": st.column_config.LinkColumn("ORCID 🔗", display_text="Ver perfil"),
                },
                hide_index=True
            )
        else:
            st.info("No se encontraron pares de coautoría. Prueba ampliando el rango de años.")

    # ── EXPORTAR ──────────────────────────────────────────
    st.markdown("---")
    st.markdown("## 📥 Exportar resultados")
    periodo_str = f"{anio_ini}–{anio_fin}" if anio_ini and anio_fin else "período completo"

    col_dl1, col_dl2, col_dl3 = st.columns(3)

    with col_dl1:
        with st.spinner("Generando PDF…"):
            pdf_bytes = generar_pdf(url_usada, df_final, res, periodo_str)
        st.download_button(
            label=f"⬇ Descargar PDF · {periodo_str}",
            data=pdf_bytes,
            file_name=f"informe_ojs_v5_{anio_ini or 'all'}_{anio_fin or 'all'}_{time.strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            type="primary",
            use_container_width=True,
        )

    with col_dl2:
        with st.spinner("Generando Excel…"):
            excel_bytes = generar_excel(df_final, res, periodo_str)
        st.download_button(
            label=f"Descargar Excel · {periodo_str}",
            data=excel_bytes,
            file_name=f"datos_ojs_v5_{anio_ini or 'all'}_{anio_fin or 'all'}_{time.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_dl3:
        csv_data = df_final.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Descargar CSV",
            data=csv_data,
            file_name=f"articulos_ojs_{anio_ini or 'all'}_{anio_fin or 'all'}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    st.info(f"Exportaciones listas · {len(df_final)} artículos · Período {periodo_str} · "
            f"Excel con {10} hojas temáticas")


elif st.session_state.df_issues is None:
    st.markdown(f"""
    <div class="info-box">
        <h4> ¿Cómo empezar?</h4>
        <p>Ingresa la URL del <b>archivo de números</b> de una revista OJS en el panel lateral
        y presiona <b>Cargar revista</b>.</p>
        <p><b>Ejemplo:</b> <code>https://revistagi.geofisica.unam.mx/index.php/RGI/issue/archive</code></p>
        <p><b>Novedades v5:</b></p>
        <ul>
          <li> Paleta de colores azul personalizada (texto en negro)</li>
          <li>Excel con <b>10 hojas temáticas</b> separadas (artículos, autores, clusters, tópicos, red...)</li>
          <li>PDF con figuras + <b>explicaciones detalladas</b> de cada análisis</li>
          <li>Filtro ampliado de stopwords académicas para clusters más precisos</li>
          <li>Tópico regional detectado automáticamente en LDA</li>
          <li>Red de coautoría corregida y más robusta</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)